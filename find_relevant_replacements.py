"""
Find relevant ASBL blog links to replace irrelevant internal links.

For each blog that has irrelevant internal links, this script finds the same
number of relevant blog posts from the ASBL catalog that are topically related
to the host blog using TF-IDF weighted similarity.

Input:  irrelevant_links.xlsx (from analyze_irrelevant_links.py)
Output: relevant_replacements.xlsx

Run:
    python find_relevant_replacements.py
    python find_relevant_replacements.py --workers 10 --out relevant_replacements.xlsx
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urlparse

import pandas as pd
from bs4 import BeautifulSoup
from tqdm import tqdm

from scrape_asbl_blogs import (
    BlogListing,
    _clean_text,
    collect_listings,
    fetch,
    make_session,
)

STOP_WORDS: Set[str] = {
    "a", "an", "the", "and", "or", "but", "in", "on", "at", "to", "for", "of",
    "is", "are", "was", "were", "be", "been", "being", "have", "has", "had",
    "do", "does", "did", "will", "would", "could", "should", "may", "might",
    "must", "shall", "can", "need", "your", "you", "our", "their", "its", "it",
    "this", "that", "these", "those", "with", "from", "by", "as", "about", "into",
    "through", "during", "before", "after", "above", "below", "between", "under",
    "again", "further", "then", "once", "here", "there", "when", "where", "why",
    "how", "all", "each", "few", "more", "most", "other", "some", "such", "no",
    "nor", "not", "only", "own", "same", "so", "than", "too", "very", "just",
    "also", "now", "get", "blog", "asbl", "https", "http", "www", "com",
    "source", "related", "guide", "article", "html", "utm",
    "read", "click", "here", "learn", "see", "know", "what", "which", "who",
    "make", "made", "like", "one", "two", "don", "every",
}


@dataclass
class BlogCatalogEntry:
    """A blog post in the ASBL catalog with its topic data."""
    url: str
    title: str
    title_tokens: List[str] = field(default_factory=list)
    body_tokens: List[str] = field(default_factory=list)
    all_tokens: List[str] = field(default_factory=list)
    slug_tokens: List[str] = field(default_factory=list)


@dataclass
class ReplacementRow:
    blog_name: str
    blog_link: str
    irrelevant_link_url: str
    irrelevant_link_text: str
    replacement_url: str
    replacement_title: str
    match_score: float
    shared_keywords: str


def _tokenize(text: str) -> List[str]:
    words = re.findall(r"[a-z0-9]+", (text or "").lower())
    return [w for w in words if len(w) > 2 and w not in STOP_WORDS]


def _slug_tokens(url: str) -> List[str]:
    path = urlparse(url).path
    slug = path.rstrip("/").split("/")[-1]
    slug = re.sub(r"[%₹\-]", " ", slug)
    return _tokenize(slug)


def _extract_body_text(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    el = soup.select_one(".entry-content, .post_content")
    if el:
        return _clean_text(el.get_text())
    return ""


def _extract_title(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    meta = soup.find("meta", property="og:title")
    if meta and meta.get("content"):
        return _clean_text(meta["content"])
    h1 = soup.find("h1", class_="entry-title") or soup.find("h1")
    if h1:
        return _clean_text(h1.get_text())
    return ""


def scrape_catalog_entry(session, listing: BlogListing) -> Optional[BlogCatalogEntry]:
    """Fetch a blog post and extract its topic tokens."""
    html = fetch(session, listing.url)
    if not html:
        return None

    title = _extract_title(html) or listing.title or ""
    body = _extract_body_text(html)

    title_tokens = _tokenize(title)
    slug_tokens = _slug_tokens(listing.url)
    body_tokens = _tokenize(body[:4000])

    all_tokens = title_tokens + slug_tokens + body_tokens

    return BlogCatalogEntry(
        url=listing.url,
        title=title,
        title_tokens=title_tokens,
        body_tokens=body_tokens,
        all_tokens=all_tokens,
        slug_tokens=slug_tokens,
    )


class TFIDFMatcher:
    """TF-IDF based blog matcher that down-weights common terms."""

    def __init__(self, catalog: List[BlogCatalogEntry]):
        self.catalog = catalog
        self.n_docs = len(catalog)
        self.idf: Dict[str, float] = {}
        self.doc_vectors: List[Dict[str, float]] = []
        self._build_index()

    def _build_index(self):
        doc_freq: Counter = Counter()
        for entry in self.catalog:
            unique_tokens = set(entry.all_tokens)
            for token in unique_tokens:
                doc_freq[token] += 1

        for token, df in doc_freq.items():
            self.idf[token] = math.log(self.n_docs / (1 + df))

        for entry in self.catalog:
            tf = Counter(entry.all_tokens)
            max_tf = max(tf.values()) if tf else 1
            vector: Dict[str, float] = {}
            for token, count in tf.items():
                normalized_tf = 0.5 + 0.5 * (count / max_tf)
                vector[token] = normalized_tf * self.idf.get(token, 0)
            self.doc_vectors.append(vector)

    def _get_query_vector(self, tokens: List[str], title_tokens: List[str]) -> Dict[str, float]:
        """Build a query vector with title tokens weighted 3x."""
        boosted_tokens = title_tokens * 3 + tokens
        tf = Counter(boosted_tokens)
        max_tf = max(tf.values()) if tf else 1
        vector: Dict[str, float] = {}
        for token, count in tf.items():
            normalized_tf = 0.5 + 0.5 * (count / max_tf)
            vector[token] = normalized_tf * self.idf.get(token, 0)
        return vector

    def _cosine_similarity(self, vec_a: Dict[str, float], vec_b: Dict[str, float]) -> float:
        common_keys = set(vec_a.keys()) & set(vec_b.keys())
        if not common_keys:
            return 0.0

        dot_product = sum(vec_a[k] * vec_b[k] for k in common_keys)
        norm_a = math.sqrt(sum(v * v for v in vec_a.values()))
        norm_b = math.sqrt(sum(v * v for v in vec_b.values()))

        if norm_a == 0 or norm_b == 0:
            return 0.0
        return dot_product / (norm_a * norm_b)

    def _get_top_shared_keywords(
        self, query_vec: Dict[str, float], doc_vec: Dict[str, float], top_n: int = 6
    ) -> List[str]:
        """Get the most important shared keywords based on combined TF-IDF weight."""
        common = set(query_vec.keys()) & set(doc_vec.keys())
        scored = [(k, query_vec[k] * doc_vec[k]) for k in common]
        scored.sort(key=lambda x: -x[1])
        return [k for k, _ in scored[:top_n]]

    def find_best_matches(
        self,
        blog_entry: Optional[BlogCatalogEntry],
        blog_url: str,
        blog_title: str,
        num_needed: int,
        exclude_urls: Set[str],
        global_usage: Counter,
        max_global_usage: int = 5,
    ) -> List[Tuple[BlogCatalogEntry, float, List[str]]]:
        """Find the top N relevant blogs, ensuring diversity."""
        if blog_entry:
            query_tokens = blog_entry.all_tokens
            title_tokens = blog_entry.title_tokens
        else:
            title_tokens = _tokenize(blog_title)
            query_tokens = title_tokens + _slug_tokens(blog_url)

        query_vec = self._get_query_vector(query_tokens, title_tokens)
        blog_url_normalized = blog_url.rstrip("/").lower()

        scored: List[Tuple[int, float, List[str]]] = []
        for i, entry in enumerate(self.catalog):
            entry_url_normalized = entry.url.rstrip("/").lower()
            if entry_url_normalized == blog_url_normalized:
                continue
            if entry_url_normalized in exclude_urls:
                continue

            sim = self._cosine_similarity(query_vec, self.doc_vectors[i])
            if sim > 0.01:
                keywords = self._get_top_shared_keywords(query_vec, self.doc_vectors[i])
                scored.append((i, sim, keywords))

        scored.sort(key=lambda x: -x[1])

        results: List[Tuple[BlogCatalogEntry, float, List[str]]] = []
        used_in_this_blog: Set[str] = set()

        for idx, sim, keywords in scored:
            if len(results) >= num_needed:
                break
            entry = self.catalog[idx]
            entry_url_normalized = entry.url.rstrip("/").lower()

            if entry_url_normalized in used_in_this_blog:
                continue
            if global_usage[entry_url_normalized] >= max_global_usage:
                continue

            results.append((entry, sim, keywords))
            used_in_this_blog.add(entry_url_normalized)
            global_usage[entry_url_normalized] += 1

        if len(results) < num_needed:
            for idx, sim, keywords in scored:
                if len(results) >= num_needed:
                    break
                entry = self.catalog[idx]
                entry_url_normalized = entry.url.rstrip("/").lower()
                if entry_url_normalized in used_in_this_blog:
                    continue
                results.append((entry, sim, keywords))
                used_in_this_blog.add(entry_url_normalized)
                global_usage[entry_url_normalized] += 1

        return results


def load_irrelevant_internal_links(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Irrelevant Links")
    internal = df[df["Link Type"] == "Internal"].copy()
    return internal


def write_replacements_excel(rows: List[ReplacementRow], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Relevant Replacements"

    headers = [
        "Blog Name",
        "Blog Link",
        "Irrelevant Link Text",
        "Irrelevant Link URL",
        "Suggested Replacement URL",
        "Suggested Replacement Title",
        "Match Score",
        "Shared Keywords",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E7D32")
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")
    good_fill = PatternFill("solid", fgColor="C8E6C9")
    irr_fill = PatternFill("solid", fgColor="FFC7CE")

    for row_data in rows:
        row_idx = ws.max_row + 1

        ws.cell(row=row_idx, column=1, value=row_data.blog_name).alignment = wrap

        blog_cell = ws.cell(row=row_idx, column=2, value=row_data.blog_link)
        blog_cell.hyperlink = row_data.blog_link
        blog_cell.font = link_font
        blog_cell.alignment = wrap

        ws.cell(row=row_idx, column=3, value=row_data.irrelevant_link_text).alignment = wrap

        irr_cell = ws.cell(row=row_idx, column=4, value=row_data.irrelevant_link_url)
        irr_cell.hyperlink = row_data.irrelevant_link_url
        irr_cell.font = link_font
        irr_cell.alignment = wrap
        irr_cell.fill = irr_fill

        rep_cell = ws.cell(row=row_idx, column=5, value=row_data.replacement_url)
        if row_data.replacement_url.startswith("http"):
            rep_cell.hyperlink = row_data.replacement_url
            rep_cell.fill = good_fill
        rep_cell.font = link_font
        rep_cell.alignment = wrap

        ws.cell(row=row_idx, column=6, value=row_data.replacement_title).alignment = wrap
        ws.cell(row=row_idx, column=7, value=row_data.match_score).alignment = wrap
        ws.cell(row=row_idx, column=8, value=row_data.shared_keywords).alignment = wrap

    widths = [45, 60, 35, 60, 60, 45, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Find relevant blog replacements for irrelevant internal links."
    )
    parser.add_argument("--input", default="irrelevant_links.xlsx", help="Input irrelevant links file.")
    parser.add_argument("--start", type=int, default=1, help="First listing page. Default: 1")
    parser.add_argument("--end", type=int, default=43, help="Last listing page. Default: 43")
    parser.add_argument("--workers", type=int, default=8, help="Parallel workers. Default: 8")
    parser.add_argument("--out", default="relevant_replacements.xlsx", help="Output Excel path.")
    parser.add_argument(
        "--max-usage", type=int, default=5,
        help="Max times a single blog can be suggested as replacement. Default: 5",
    )
    args = parser.parse_args()

    print("Step 1: Loading irrelevant internal links...", file=sys.stderr)
    irr_df = load_irrelevant_internal_links(args.input)
    print(f"  Found {len(irr_df)} irrelevant internal links across {irr_df['Blog Link'].nunique()} blogs.", file=sys.stderr)

    print("\nStep 2: Building blog catalog (scraping all ASBL blogs)...", file=sys.stderr)
    session = make_session()
    listings = collect_listings(session, args.start, args.end)
    print(f"  Found {len(listings)} blog posts in catalog.", file=sys.stderr)

    catalog: List[BlogCatalogEntry] = []
    failed: List[str] = []

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {
            pool.submit(scrape_catalog_entry, make_session(), entry): entry
            for entry in listings
        }
        for fut in tqdm(as_completed(futures), total=len(futures), desc="Building catalog"):
            entry = futures[fut]
            try:
                result = fut.result()
            except Exception as exc:
                print(f"  ! Error: {entry.url}: {exc}", file=sys.stderr)
                failed.append(entry.url)
                continue
            if result is None:
                failed.append(entry.url)
                continue
            catalog.append(result)

    print(f"  Catalog built: {len(catalog)} blogs indexed ({len(failed)} failed).", file=sys.stderr)

    print("\nStep 3: Building TF-IDF index...", file=sys.stderr)
    matcher = TFIDFMatcher(catalog)
    print("  Index built.", file=sys.stderr)

    catalog_by_url: Dict[str, BlogCatalogEntry] = {
        e.url.rstrip("/").lower(): e for e in catalog
    }

    print("\nStep 4: Finding relevant replacements...", file=sys.stderr)

    blogs_grouped = irr_df.groupby("Blog Link", sort=False)
    replacement_rows: List[ReplacementRow] = []
    global_usage: Counter = Counter()

    for blog_link, group in tqdm(blogs_grouped, desc="Matching replacements"):
        blog_name = group.iloc[0]["Blog Name"]
        blog_url_normalized = blog_link.rstrip("/").lower()

        blog_entry = catalog_by_url.get(blog_url_normalized)

        existing_internal_links: Set[str] = set()
        for _, row in group.iterrows():
            existing_internal_links.add(row["Link URL"].rstrip("/").lower())

        num_needed = len(group)

        replacements = matcher.find_best_matches(
            blog_entry=blog_entry,
            blog_url=blog_link,
            blog_title=blog_name,
            num_needed=num_needed,
            exclude_urls=existing_internal_links,
            global_usage=global_usage,
            max_global_usage=args.max_usage,
        )

        irr_rows = list(group.iterrows())
        for i, (_, irr_row) in enumerate(irr_rows):
            if i < len(replacements):
                entry, score, keywords = replacements[i]
                replacement_rows.append(ReplacementRow(
                    blog_name=blog_name,
                    blog_link=blog_link,
                    irrelevant_link_url=irr_row["Link URL"],
                    irrelevant_link_text=irr_row["Link Text"],
                    replacement_url=entry.url,
                    replacement_title=entry.title,
                    match_score=round(score, 4),
                    shared_keywords=", ".join(keywords),
                ))
            else:
                replacement_rows.append(ReplacementRow(
                    blog_name=blog_name,
                    blog_link=blog_link,
                    irrelevant_link_url=irr_row["Link URL"],
                    irrelevant_link_text=irr_row["Link Text"],
                    replacement_url="(no suitable match found)",
                    replacement_title="",
                    match_score=0.0,
                    shared_keywords="",
                ))

    print(f"\nStep 5: Writing output to {args.out}...", file=sys.stderr)
    write_replacements_excel(replacement_rows, args.out)

    matched = sum(1 for r in replacement_rows if r.replacement_url.startswith("http"))
    unmatched = len(replacement_rows) - matched
    unique_replacements = len({r.replacement_url for r in replacement_rows if r.replacement_url.startswith("http")})

    print(
        f"\nDone! Saved {args.out}\n"
        f"  Total irrelevant internal links: {len(replacement_rows)}\n"
        f"  Matched with replacements:       {matched}\n"
        f"  No suitable match found:         {unmatched}\n"
        f"  Unique replacement blogs used:   {unique_replacements}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
