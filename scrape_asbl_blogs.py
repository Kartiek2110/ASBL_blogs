"""
Scrape every blog post on https://asbl.in/blog/ across all paginated pages
(default 1..43), pull the styled `<a style="color: #0000ff;" href="...">…</a>`
links out of each post body, classify them as internal (asbl.in) vs external,
and dump everything into an Excel workbook.

Output columns (one row per blog post; extra rows when a post has multiple links):
    - Blog Name
    - Blog Link          (clickable hyperlink)
    - Blog Date
    - Internal Link      (clickable hyperlink; one per row)
    - External Link      (clickable hyperlink; one per row)

Run:
    python scrape_asbl_blogs.py
    python scrape_asbl_blogs.py --start 1 --end 43 --workers 8 --out asbl_blogs.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Optional, Tuple
from urllib.parse import urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm

BASE_URL = "https://asbl.in/blog/"
INTERNAL_HOST = "asbl.in"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Matches <a style="color: #0000ff;" href="...">label</a>
# - Tolerates extra spaces, optional trailing semicolon, attribute reordering
#   and other attributes on the anchor (target, rel, etc.).
STYLED_ANCHOR_RE = re.compile(
    r'<a\b[^>]*?style\s*=\s*"[^"]*color\s*:\s*#0000ff[^"]*"[^>]*?>(.*?)</a>',
    re.IGNORECASE | re.DOTALL,
)


@dataclass
class BlogListing:
    """A blog post entry as advertised on a listing page."""

    url: str
    title: str = ""
    date: str = ""


@dataclass
class BlogRecord:
    name: str
    link: str
    date: str
    internal_links: List[Tuple[str, str]] = field(default_factory=list)  # (label, url)
    external_links: List[Tuple[str, str]] = field(default_factory=list)


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def fetch(session: requests.Session, url: str, retries: int = 3, timeout: int = 20) -> Optional[str]:
    """GET a URL with simple exponential backoff. Returns text or None on hard failure."""
    delay = 1.0
    for attempt in range(1, retries + 1):
        try:
            r = session.get(url, timeout=timeout)
            if r.status_code == 200:
                return r.text
            if r.status_code == 404:
                return None
            # Retry on 5xx / 429
            if r.status_code in (429, 500, 502, 503, 504):
                time.sleep(delay)
                delay *= 2
                continue
            return None
        except requests.RequestException:
            time.sleep(delay)
            delay *= 2
    return None


def parse_listing_page(html: str) -> List[BlogListing]:
    """Return blog post listings (url, title, date) from a paginated listing page.

    The ASBL listing renders each post inside `.insight-content` like:

        <div class="insight-content">
            <span class="insight-cat">
                <div class="insight-date">April 8, 2026</div>
                <a href="…category…">Design</a>
            </span>
            <h3 class="post-title"><a href="POST_URL">Title</a></h3>
            ...
        </div>

    But many cards instead wrap the *entire* card in an `<a>` (so the title
    `h3.post-title` contains no anchor). Handle both shapes.
    """
    soup = BeautifulSoup(html, "lxml")
    listings: List[BlogListing] = []
    seen: set = set()

    for h3 in soup.select("h3.post-title"):
        # 1. Title comes from h3 text.
        title = _clean_text(h3.get_text())

        # 2. URL: inner anchor first, otherwise walk up to find a wrapping anchor.
        href: Optional[str] = None
        inner_a = h3.find("a", href=True)
        if inner_a:
            href = inner_a["href"].strip()
        else:
            parent = h3.parent
            while parent is not None and parent.name != "[document]":
                if parent.name == "a" and parent.get("href"):
                    href = parent["href"].strip()
                    break
                parent = parent.parent

        if not href:
            continue
        if not href.startswith("https://asbl.in/blog/"):
            continue
        if "/page/" in href or href.rstrip("/") == "https://asbl.in/blog":
            continue
        if href in seen:
            continue
        seen.add(href)

        # 3. Date: look for sibling `.insight-date` inside the enclosing
        #    `.insight-content` container.
        date = ""
        container = h3.find_parent(class_="insight-content") or h3.find_parent("article")
        if container:
            date_el = container.select_one(".insight-date, .posted-on, .entry-date")
            if date_el:
                date = _clean_text(date_el.get_text())

        listings.append(BlogListing(url=href, title=title, date=date))

    return listings


def _clean_text(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


def _extract_date(soup: BeautifulSoup, html: str) -> str:
    """Try several strategies to pull the published date from a blog post."""

    # 1. <meta property="article:published_time" content="...">
    meta = soup.find("meta", property="article:published_time")
    if meta and meta.get("content"):
        return _format_iso_date(meta["content"])

    # 2. <time datetime="...">
    t = soup.find("time")
    if t and t.get("datetime"):
        return _format_iso_date(t["datetime"])
    if t and _clean_text(t.get_text()):
        return _clean_text(t.get_text())

    # 3. AIOSEO schema JSON-LD often carries a datePublished field.
    m = re.search(r'"datePublished"\s*:\s*"([^"]+)"', html)
    if m:
        return _format_iso_date(m.group(1))

    # 4. .posted-on / .entry-date selectors
    el = soup.select_one(".posted-on, .entry-date, .ast-post-date")
    if el:
        return _clean_text(el.get_text())

    return ""


def _format_iso_date(value: str) -> str:
    """Normalize an ISO-8601 timestamp to YYYY-MM-DD; otherwise return as-is."""
    try:
        # Python's fromisoformat handles offsets from 3.11+; fall back to strptime patterns.
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        try:
            dt = datetime.strptime(value[:10], "%Y-%m-%d")
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            return value


def _extract_title(soup: BeautifulSoup) -> str:
    h1 = soup.find("h1", class_="entry-title") or soup.find("h1")
    if h1 and _clean_text(h1.get_text()):
        return _clean_text(h1.get_text())

    meta = soup.find("meta", property="og:title")
    if meta and meta.get("content"):
        return _clean_text(meta["content"])

    if soup.title and soup.title.string:
        return _clean_text(soup.title.string)
    return ""


def _is_internal(url: str) -> bool:
    try:
        host = urlparse(url).netloc.lower()
    except ValueError:
        return False
    return host == INTERNAL_HOST or host.endswith("." + INTERNAL_HOST)


def _extract_styled_links(html: str) -> List[Tuple[str, str]]:
    """Find every <a style="color: #0000ff;" href="...">label</a> in the post body."""
    # Limit the search to the article body when possible to avoid template noise.
    body_match = re.search(
        r'<div[^>]+class="[^"]*(?:entry-content|post_content)[^"]*"[^>]*>(.*?)</article>',
        html,
        re.IGNORECASE | re.DOTALL,
    )
    scope = body_match.group(1) if body_match else html

    links: List[Tuple[str, str]] = []
    for m in STYLED_ANCHOR_RE.finditer(scope):
        anchor_html = m.group(0)
        label_html = m.group(1)
        href_match = re.search(r'href\s*=\s*"([^"]+)"', anchor_html, re.IGNORECASE)
        if not href_match:
            continue
        href = href_match.group(1).strip()
        label = _clean_text(BeautifulSoup(label_html, "lxml").get_text())
        links.append((label, href))
    return links


def parse_blog_post(url: str, html: str) -> BlogRecord:
    soup = BeautifulSoup(html, "lxml")
    rec = BlogRecord(
        name=_extract_title(soup),
        link=url,
        date=_extract_date(soup, html),
    )
    for label, href in _extract_styled_links(html):
        if _is_internal(href):
            rec.internal_links.append((label, href))
        else:
            rec.external_links.append((label, href))
    return rec


def collect_listings(session: requests.Session, start: int, end: int) -> List[BlogListing]:
    listings: List[BlogListing] = []
    seen: set = set()
    print(f"Collecting post URLs from pages {start}..{end}", file=sys.stderr)
    for page in tqdm(range(start, end + 1), desc="Listing pages"):
        page_url = BASE_URL if page == 1 else f"{BASE_URL}page/{page}/"
        html = fetch(session, page_url)
        if not html:
            print(f"  ! Failed to fetch listing page {page}: {page_url}", file=sys.stderr)
            continue
        for entry in parse_listing_page(html):
            if entry.url in seen:
                continue
            seen.add(entry.url)
            listings.append(entry)
    return listings


def scrape_post(session: requests.Session, listing: BlogListing) -> Optional[BlogRecord]:
    """Fetch and parse a single blog post, using listing metadata as a fallback."""
    html = fetch(session, listing.url)
    if not html:
        return None
    rec = parse_blog_post(listing.url, html)
    if not rec.name and listing.title:
        rec.name = listing.title
    if not rec.date and listing.date:
        rec.date = listing.date
    return rec


def _link_display(label: str, href: str) -> str:
    clean = (label or "").strip()
    if clean.startswith("[") and clean.endswith("]"):
        clean = clean[1:-1].strip()
    return clean or href


def _set_hyperlink(cell, url: str, display: str, link_font) -> None:
    cell.value = display
    cell.hyperlink = url
    cell.font = link_font


def _parse_link_lines(text: object) -> List[Tuple[str, str]]:
    """Parse 'label: url' lines (or bare URLs) from an existing Excel cell."""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    links: List[Tuple[str, str]] = []
    for line in str(text).splitlines():
        line = line.strip()
        if not line:
            continue
        if ": http" in line:
            label, _, url = line.partition(": ")
            links.append((label.strip(), url.strip()))
        elif line.startswith("http"):
            links.append(("", line))
    return links


def records_from_excel(path: str) -> List[BlogRecord]:
    """Rebuild BlogRecord objects from an existing workbook export."""
    df = pd.read_excel(path)
    records: List[BlogRecord] = []
    current: Optional[BlogRecord] = None

    for _, row in df.iterrows():
        name = row.get("Blog Name", "")
        link = row.get("Blog Link", "")
        date = row.get("Blog Date", "")

        name = "" if pd.isna(name) else str(name).strip()
        link = "" if pd.isna(link) else str(link).strip()
        date = "" if pd.isna(date) else str(date).strip()
        if date.endswith(" 00:00:00"):
            date = date[:10]

        internal_col = row.get("Internal Link", row.get("Internal Links", ""))
        external_col = row.get("External Link", row.get("External Links", ""))
        internal = _parse_link_lines(internal_col)
        external = _parse_link_lines(external_col)

        if link.startswith("http"):
            current = BlogRecord(name=name, link=link, date=date)
            current.internal_links.extend(internal)
            current.external_links.extend(external)
            records.append(current)
        elif current is not None and (internal or external):
            current.internal_links.extend(internal)
            current.external_links.extend(external)

    return records


def write_excel(records: List[BlogRecord], path: str) -> None:
    """Write records to Excel with clickable hyperlinks on every URL."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = [
        "Blog Name",
        "Blog Link",
        "Blog Date",
        "Internal Link",
        "External Link",
        "# Internal",
        "# External",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "ASBL Blogs"
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")
    row_idx = 2

    for rec in records:
        link_rows = max(len(rec.internal_links), len(rec.external_links), 1)
        start_row = row_idx

        for i in range(link_rows):
            internal = rec.internal_links[i] if i < len(rec.internal_links) else None
            external = rec.external_links[i] if i < len(rec.external_links) else None

            ws.cell(row=row_idx, column=1, value=rec.name).alignment = wrap
            _set_hyperlink(
                ws.cell(row=row_idx, column=2),
                rec.link,
                rec.link,
                link_font,
            )
            ws.cell(row=row_idx, column=2).alignment = wrap

            ws.cell(row=row_idx, column=3, value=rec.date).alignment = wrap

            if internal:
                label, href = internal
                _set_hyperlink(
                    ws.cell(row=row_idx, column=4),
                    href,
                    _link_display(label, href),
                    link_font,
                )
            ws.cell(row=row_idx, column=4).alignment = wrap

            if external:
                label, href = external
                _set_hyperlink(
                    ws.cell(row=row_idx, column=5),
                    href,
                    _link_display(label, href),
                    link_font,
                )
            ws.cell(row=row_idx, column=5).alignment = wrap

            if i == 0:
                ws.cell(row=row_idx, column=6, value=len(rec.internal_links)).alignment = wrap
                ws.cell(row=row_idx, column=7, value=len(rec.external_links)).alignment = wrap

            row_idx += 1

        if link_rows > 1:
            for col in (1, 2, 3, 6, 7):
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=row_idx - 1,
                    end_column=col,
                )
                ws.cell(row=start_row, column=col).alignment = wrap

    widths = {
        "Blog Name": 50,
        "Blog Link": 60,
        "Blog Date": 14,
        "Internal Link": 55,
        "External Link": 55,
        "# Internal": 12,
        "# External": 12,
    }
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Scrape ASBL blog posts into an Excel workbook.")
    parser.add_argument("--start", type=int, default=1, help="First listing page (inclusive). Default: 1")
    parser.add_argument("--end", type=int, default=43, help="Last listing page (inclusive). Default: 43")
    parser.add_argument("--workers", type=int, default=8, help="Parallel post fetchers. Default: 8")
    parser.add_argument("--out", default="asbl_blogs.xlsx", help="Output Excel file path.")
    parser.add_argument("--limit", type=int, default=0, help="Optional cap on number of posts (debugging).")
    parser.add_argument(
        "--from-excel",
        default="",
        help="Skip scraping; rebuild clickable links from an existing Excel export.",
    )
    args = parser.parse_args()

    if args.from_excel:
        records = records_from_excel(args.from_excel)
        write_excel(records, args.out)
        print(f"Rewrote {len(records)} blog rows with clickable links -> {args.out}", file=sys.stderr)
        return 0

    session = make_session()
    listings = collect_listings(session, args.start, args.end)

    if args.limit:
        listings = listings[: args.limit]

    print(f"Found {len(listings)} unique blog posts.", file=sys.stderr)

    records: List[BlogRecord] = []
    failed: List[str] = []

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        future_to_listing = {
            pool.submit(scrape_post, make_session(), entry): entry for entry in listings
        }
        for fut in tqdm(as_completed(future_to_listing), total=len(future_to_listing), desc="Blog posts"):
            entry = future_to_listing[fut]
            try:
                rec = fut.result()
            except Exception as exc:  # noqa: BLE001
                print(f"  ! Error scraping {entry.url}: {exc}", file=sys.stderr)
                failed.append(entry.url)
                continue
            if rec is None:
                failed.append(entry.url)
                continue
            records.append(rec)

    # Preserve listing order (newest first) for the output.
    order = {entry.url: i for i, entry in enumerate(listings)}
    records.sort(key=lambda r: order.get(r.link, 1_000_000))

    write_excel(records, args.out)

    print(
        f"Saved {len(records)} rows to {args.out}. "
        f"({len(failed)} URLs failed to fetch.)",
        file=sys.stderr,
    )
    if failed:
        print("Failed URLs:", file=sys.stderr)
        for u in failed:
            print(f"  - {u}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
