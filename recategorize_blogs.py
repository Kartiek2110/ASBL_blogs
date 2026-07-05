"""
Re-categorize ASBL blog posts into editorial topic clusters.

Primary categories (from editorial brief):
  - Home Buying Guides
  - NRI Property Investment
  - Hyderabad Real Estate Market
  - Area Guides (Hyderabad neighborhoods)
  - Property Investment Tips

Posts that do not fit receive a new, descriptive category that matches
the blog content (e.g. Home Design & Vastu, Home Loans & Finance).

Run:
    python recategorize_blogs.py
    python recategorize_blogs.py --in asbl_post_sitemap_clusters.xlsx --out asbl_post_sitemap_clusters.xlsx
"""

from __future__ import annotations

import argparse
import html
import re
import sys
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PRIMARY_CATEGORIES = [
    "Home Buying Guides",
    "NRI Property Investment",
    "Hyderabad Real Estate Market",
    "Area Guides (Hyderabad neighborhoods)",
    "Property Investment Tips",
]

# Ordered rules: first match wins. Each rule is (category, patterns).
CLASSIFICATION_RULES: List[Tuple[str, Sequence[str]]] = [
    (
        "NRI Property Investment",
        [
            r"\bnri\b",
            r"\bnris\b",
            r"non[- ]resident indian",
            r"for nrIs\b",
            r"repatriat",
            r"currency exchange rate.{0,20}nri",
        ],
    ),
    (
        "Real Estate News & Industry Updates",
        [
            r"largest single land deal",
            r"land deal",
            r"crore/acre auction",
            r"asbl in news",
            r"industry update",
            r"breaking news",
            r"policy announcement",
        ],
    ),
    (
        "Area Guides (Hyderabad neighborhoods)",
        [
            r"\bgachibowli\b",
            r"\bkukatpally\b",
            r"\bfinancial district\b",
            r"\bpocharam\b",
            r"\bkondapur\b",
            r"\bmadhapur\b",
            r"\bmiyapur\b",
            r"\bnarsingi\b",
            r"\bkokapet\b",
            r"\btellapur\b",
            r"\bmanikonda\b",
            r"\bnanakramguda\b",
            r"\braidurg\b",
            r"\bhitec city\b",
            r"\bbachupally\b",
            r"\balwal\b",
            r"\buppal\b",
            r"\bkompally\b",
            r"\bshamshabad\b",
            r"\bbegumpet\b",
            r"\bjubilee hills\b",
            r"\bbanjara hills\b",
            r"\bsecunderabad\b",
            r"\blb nagar\b",
            r"\bdilsukhnagar\b",
            r"\btarnaka\b",
            r"\bhafeezpet\b",
            r"\bameerpet\b",
            r"\beast hyderabad\b",
            r"\bwest hyderabad\b",
            r"\bnorth hyderabad\b",
            r"\bsouth hyderabad\b",
            r"\by junction\b",
            r"\bgachibowli vs\b",
            r"\bis uppal\b",
            r"\blook east policy\b",
        ],
    ),
    (
        "Construction & Building Technology",
        [
            r"\bmivan\b",
            r"\brcc shear\b",
            r"generative ai",
            r"ai[- ]driven construction",
            r"construction processes",
            r"quality construction",
            r"construction tech",
            r"architecture with ai",
            r"streamlining construction",
        ],
    ),
    (
        "Home Loans & Finance",
        [
            r"home loan",
            r"\bcibil\b",
            r"\bemi\b",
            r"interest rate",
            r"loan approval",
            r"loan tenure",
            r"mortgage",
            r"home loan",
            r"loan sanctioned",
            r"principal.{0,15}interest",
            r"tax benefit.{0,20}home loan",
            r"pre[- ]?emi",
        ],
    ),
    (
        "Legal & Property Documentation",
        [
            r"stamp duty",
            r"registration charges",
            r"sale deed",
            r"allotment letter",
            r"agreement to sell",
            r"\btdr\b",
            r"\brera\b",
            r"deed of declaration",
            r"encumbrance certificate",
            r"property documents",
            r"legal tips",
            r"documentation",
            r"occupancy certificate",
            r"khata\b",
            r"mutation\b",
            r"conveyance deed",
            r"power of attorney",
            r"gift deed",
        ],
    ),
    (
        "Ready to Move In Homes",
        [
            r"ready to move",
            r"ready-to-move",
            r"\brtm\b",
            r"immediate possession",
            r"don't pay gst or registration",
        ],
    ),
    (
        "Home Design & Vastu",
        [
            r"\bvastu\b",
            r"interior design",
            r"design trends",
            r"design and decor",
            r"house facing",
            r"facing direction",
            r"mirror direction",
            r"colour as per",
            r"color as per",
            r"room guide",
            r"balcony",
            r"kitchen colour",
            r"kitchen color",
            r"outdoor living",
            r"house plan(?!.*invest)",
            r"flat design",
            r"3bhk flat design",
            r"2bhk interior",
            r"floor plan",
            r"architecture house plans",
            r"green building",
            r"modern architecture",
        ],
    ),
    (
        "City & Market Comparisons",
        [
            r"hyderabad vs",
            r"vs bengaluru",
            r"vs bangalore",
            r"vs mumbai",
            r"vs pune",
            r"vs chennai",
            r"best cities in india",
            r"choosing the ideal real estate market",
            r"city comparison",
            r"bengaluru vs",
        ],
    ),
    (
        "Gated Communities & Lifestyle",
        [
            r"gated community",
            r"gated communities",
            r"creche facility",
            r"kid[- ]friendly",
            r"amenities in your",
            r"community lifestyle",
            r"living in a gated",
            r"clubhouse",
            r"working parents",
        ],
    ),
    (
        "Property Investment Tips",
        [
            r"investment tip",
            r"invest in property",
            r"investing in",
            r"worth investing",
            r"\broi\b",
            r"return on investment",
            r"rental income",
            r"rental yield",
            r"appreciation",
            r"wealth fastest",
            r"flip",
            r"flips, rentals",
            r"portfolio",
            r"investment strateg",
            r"real estate investment",
            r"property investment",
            r"investment guide",
            r"investment in india",
            r"investment in hyderabad",
            r"invest now",
            r"high[- ]growth",
            r"long[- ]term benefits and roi",
            r"buy, rent, or flip",
            r"sez.{0,20}investment",
        ],
    ),
    (
        "Hyderabad Real Estate Market",
        [
            r"hyderabad real estate market",
            r"hyderabad market",
            r"hyderabad real estate 20",
            r"hyderabad property market",
            r"hyderabad housing market",
            r"hyderabad real estate demand",
            r"hyderabad real estate growth",
            r"hyderabad real estate trend",
            r"hyderabad real estate outlook",
            r"hyderabad homeownership",
            r"hyderabad real estate:",
            r"real estate in hyderabad",
            r"hyderabad's real estate",
            r"hyderabad property prices",
            r"hyderabad residential market",
            r"market in hyderabad",
            r"hyderabad.{0,30}(demand|growth|surge|boom|outlook|trend)",
            r"(demand|growth|surge|boom|outlook|trend).{0,30}hyderabad",
            r"quality of life in hyderabad",
            r"cost of living in hyderabad",
            r"residential zoning.{0,20}hyderabad",
            r"high[- ]rise apartments in hyderabad",
            r"policy changes.{0,30}hyderabad",
            r"rental market 20",
            r"shadow inventory",
        ],
    ),
    (
        "Home Maintenance & Upkeep",
        [
            r"maintenance",
            r"water leakage",
            r"home repair",
            r"upkeep",
            r"fix water",
            r"common home maintenance",
        ],
    ),
    (
        "Home Buying Guides",
        [
            r"home buying",
            r"homebuyer",
            r"home buyer",
            r"flat buyer",
            r"first[- ]time",
            r"first time",
            r"buying a flat",
            r"buying a home",
            r"buying an apartment",
            r"buy a flat",
            r"buy a home",
            r"before buying",
            r"when buying",
            r"flat inspection",
            r"property inspection",
            r"checklist",
            r"mistakes to avoid",
            r"what to check",
            r"what to look for",
            r"how to choose",
            r"how to select",
            r"how to study",
            r"how to evaluate",
            r"which floor is best",
            r"which floor",
            r"down payment",
            r"\bbhk\b",
            r"2bhk or 3bhk",
            r"3bhk or 4bhk",
            r"carpet area",
            r"built[- ]up area",
            r"super built",
            r"flat facing",
            r"deciding flat facing",
            r"booking amount",
            r"clp payment",
            r"under construction apartment",
            r"questions to ask",
            r"guide for.{0,20}buyer",
            r"buyers guide",
            r"buyer's guide",
            r"for flat buyers",
            r"for homebuyers",
            r"for home buyers",
            r"for first",
            r"hidden challenges",
            r"hidden costs",
            r"real cost of buying",
            r"gst on residential",
            r"builder charging",
            r"water[- ]resistant apartments",
            r"choosing water",
            r"flood[- ]prone hyderabad",
            r"mep drawings",
            r"select the right home",
            r"selecting the right",
            r"choose the best home",
            r"choose the right",
            r"right time to buy",
            r"smart choice",
            r"ultimate guide",
            r"simple guide",
            r"complete guide",
            r"practical guide",
            r"step[- ]by[- ]step",
        ],
    ),
]

# Fallback categories based on the legacy WordPress topic when nothing else matches.
LEGACY_CATEGORY_MAP = {
    "Finance": "Home Loans & Finance",
    "Design and Decor": "Home Design & Vastu",
    "Design": "Home Design & Vastu",
    "Construction Tech": "Construction & Building Technology",
    "Gated Communities": "Gated Communities & Lifestyle",
    "ready to move in": "Ready to Move In Homes",
    "ASBL in News": "Real Estate News & Industry Updates",
    "Hyderabad's Residential locations": "Area Guides (Hyderabad neighborhoods)",
    "Real Estate 1O1": "Home Buying Guides",
    "Hyderabad Real Estate": "Hyderabad Real Estate Market",
}


def _clean_title(title: object) -> str:
    text = html.unescape(str(title or ""))
    return re.sub(r"<[^>]+>", "", text).strip()


def _search_text(row: pd.Series) -> str:
    title = _clean_title(row.get("Blog Title", ""))
    slug = str(row.get("Slug", "") or "")
    old_category = str(row.get("Topic Category", "") or "")
    subcategories = str(row.get("Subcategories", "") or "")
    return " ".join([title, slug, old_category, subcategories]).lower()


def classify_blog(row: pd.Series) -> Tuple[str, str]:
    """Return (new_category, match_reason)."""
    text = _search_text(row)

    for category, patterns in CLASSIFICATION_RULES:
        for pattern in patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return category, f"rule:{pattern}"

    legacy = str(row.get("Topic Category", "") or "").strip()
    if legacy in LEGACY_CATEGORY_MAP:
        return LEGACY_CATEGORY_MAP[legacy], f"legacy:{legacy}"

    if "hyderabad" in text:
        return "Hyderabad Real Estate Market", "fallback:hyderabad"

    if any(word in text for word in ("invest", "roi", "rental", "appreciation")):
        return "Property Investment Tips", "fallback:investment"

    if any(word in text for word in ("buy", "buyer", "bhk", "flat", "apartment", "home")):
        return "Home Buying Guides", "fallback:buying"

    return "General Real Estate Guides", "fallback:general"


def classify_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    results = [classify_blog(row) for _, row in out.iterrows()]
    out["Editorial Category"] = [cat for cat, _ in results]
    out["Classification Notes"] = [note for _, note in results]
    return out


def _sheet_name(category: str, used: Dict[str, int]) -> str:
    base = re.sub(r'[\[\]\*\?:/\\]', "", category).strip() or "Uncategorized"
    base = base[:31]
    if base not in used:
        used[base] = 0
        return base
    used[base] += 1
    suffix = f" ({used[base]})"
    return f"{base[: 31 - len(suffix)]}{suffix}"


def write_clustered_excel(df: pd.DataFrame, path: str) -> None:
    """Write a multi-sheet workbook grouped by Editorial Category."""
    from openpyxl import Workbook

    # Drop old per-category sheets content; rebuild from classified data.
    category_col = "Editorial Category"
    headers = [c for c in df.columns if c != "Classification Notes"] + []

    # Preferred column order
    preferred = [
        "Blog Title",
        "URL",
        "Slug",
        "Editorial Category",
        "Topic Category",
        "Subcategories",
        "Images",
        "Change Frequency",
        "Priority",
        "Priority Label",
        "Last Updated",
        "Last Updated Date",
        "Last Updated Time",
        "Published Date",
    ]
    headers = [h for h in preferred if h in df.columns]
    extra = [h for h in df.columns if h not in headers and h != "Classification Notes"]
    headers.extend(extra)

    counts: Dict[str, int] = {}
    for cat in df[category_col]:
        counts[str(cat)] = counts.get(str(cat), 0) + 1

    grouped = df.copy()
    grouped["_sort_count"] = grouped[category_col].map(lambda c: -counts.get(str(c), 0))
    grouped = grouped.sort_values(
        by=["_sort_count", category_col, "Blog Title"],
        key=lambda col: col.str.lower() if col.name in (category_col, "Blog Title") else col,
    ).drop(columns="_sort_count")

    wb = Workbook()
    used_names: Dict[str, int] = {}
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")

    def write_sheet(ws, frame: pd.DataFrame) -> None:
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for _, row in frame.iterrows():
            values = []
            for header in headers:
                val = row.get(header, "")
                if pd.isna(val):
                    val = ""
                values.append(val)
            ws.append(values)
            row_idx = ws.max_row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = wrap
                if header == "URL" and cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.font = link_font

        widths = {
            "Blog Title": 55,
            "URL": 65,
            "Slug": 35,
            "Editorial Category": 34,
            "Topic Category": 22,
            "Subcategories": 28,
            "Images": 10,
            "Change Frequency": 16,
            "Priority": 10,
            "Priority Label": 14,
            "Last Updated": 24,
            "Last Updated Date": 18,
            "Last Updated Time": 14,
            "Published Date": 14,
        }
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 18)
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    ws_all = wb.active
    ws_all.title = "All Posts"
    write_sheet(ws_all, grouped)

    ws_summary = wb.create_sheet("Category Summary")
    ws_summary.append(["Editorial Category", "Post Count", "Share %", "Primary Category?"])
    for col_idx in range(1, 5):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    total = len(grouped) or 1
    summary_rows = sorted(counts.items(), key=lambda item: (-item[1], item[0].lower()))
    for row_idx, (category, count) in enumerate(summary_rows, start=2):
        ws_summary.cell(row=row_idx, column=1, value=category)
        ws_summary.cell(row=row_idx, column=2, value=count)
        ws_summary.cell(row=row_idx, column=3, value=round(100 * count / total, 1))
        ws_summary.cell(
            row=row_idx,
            column=4,
            value="Yes" if category in PRIMARY_CATEGORIES else "New",
        )
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 16
    ws_summary.freeze_panes = "A2"

    clusters: Dict[str, pd.DataFrame] = {}
    for category, frame in grouped.groupby(category_col, sort=True):
        clusters[str(category)] = frame.sort_values(
            by="Blog Title", key=lambda s: s.str.lower()
        )

    for category in sorted(clusters, key=str.lower):
        sheet_name = _sheet_name(category, used_names)
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, clusters[category])

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Re-categorize ASBL blogs into editorial clusters.")
    parser.add_argument("--in", dest="input_path", default="asbl_post_sitemap_clusters.xlsx")
    parser.add_argument("--out", dest="output_path", default="asbl_post_sitemap_clusters.xlsx")
    parser.add_argument(
        "--also-update",
        default="",
        help="Optional second Excel path to update with the same categories.",
    )
    args = parser.parse_args()

    df = pd.read_excel(args.input_path, sheet_name="All Posts")
    classified = classify_dataframe(df)

    write_clustered_excel(classified, args.output_path)

    counts = classified["Editorial Category"].value_counts()
    print(f"Saved {len(classified)} posts -> {args.output_path}", file=sys.stderr)
    print("\nEditorial Category distribution:", file=sys.stderr)
    for cat, n in counts.items():
        tag = "PRIMARY" if cat in PRIMARY_CATEGORIES else "NEW"
        print(f"  [{tag}] {cat}: {n}", file=sys.stderr)

    if args.also_update:
        write_clustered_excel(classified, args.also_update)
        print(f"Also updated {args.also_update}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
