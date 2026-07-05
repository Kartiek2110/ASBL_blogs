"""
Scrape ASBL blog post metadata from post-sitemap.xml, enrich each entry with
its WordPress topic category, and export clustered results to Excel.

Data sources:
  - https://asbl.in/blog/post-sitemap.xml  (URL, images, changefreq, priority, lastmod)
  - https://asbl.in/blog/wp-json/wp/v2/posts  (title, published date, category)

Output workbook:
  - All Posts        – every sitemap row with topic cluster columns
  - Category Summary – post count per topic cluster
  - One sheet per topic cluster (e.g. Design, Hyderabad Real Estate, …)

Run:
    python scrape_post_sitemap.py
    python scrape_post_sitemap.py --out asbl_post_sitemap_clusters.xlsx
"""

from __future__ import annotations

import argparse
import html
import re
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Tuple
from urllib.parse import unquote, urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SITEMAP_URL = "https://asbl.in/blog/post-sitemap.xml"
WP_POSTS_URL = "https://asbl.in/blog/wp-json/wp/v2/posts"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

SITEMAP_NS = {
    "sm": "http://www.sitemaps.org/schemas/sitemap/0.9",
    "img": "http://www.google.com/schemas/sitemap-image/1.1",
}

PRIORITY_LABELS = {
    "1.0": "Highest",
    "0.9": "High",
    "0.8": "High",
    "0.7": "Medium",
    "0.6": "Medium",
    "0.5": "Medium",
    "0.4": "Low",
    "0.3": "Low",
    "0.2": "Low",
    "0.1": "Low",
}


@dataclass
class SitemapEntry:
    url: str
    images: int = 0
    change_frequency: str = ""
    priority: str = ""
    priority_label: str = ""
    last_updated: str = ""
    last_updated_date: str = ""
    last_updated_time: str = ""
    title: str = ""
    slug: str = ""
    topic_category: str = "Uncategorized"
    subcategories: List[str] = field(default_factory=list)
    published_date: str = ""


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def fetch_text(session: requests.Session, url: str, retries: int = 3, timeout: int = 30) -> str:
    delay = 1.0
    for _ in range(retries):
        try:
            response = session.get(url, timeout=timeout)
            if response.status_code == 200:
                return response.text
            if response.status_code in (429, 500, 502, 503, 504):
                time.sleep(delay)
                delay *= 2
                continue
            response.raise_for_status()
        except requests.RequestException:
            time.sleep(delay)
            delay *= 2
    raise RuntimeError(f"Failed to fetch {url}")


def _strip_cdata(value: str) -> str:
    return (value or "").strip()


def _format_lastmod(iso_value: str) -> Tuple[str, str, str]:
    """Return (combined, date, time) from an ISO-8601 lastmod string."""
    if not iso_value:
        return "", "", ""
    try:
        dt = datetime.fromisoformat(iso_value.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.astimezone()
        date_part = dt.strftime("%B %d, %Y")
        time_part = dt.strftime("%I:%M %p").lstrip("0")
        return f"{date_part} {time_part}", date_part, time_part
    except ValueError:
        return iso_value, iso_value, ""


def _priority_label(value: str) -> str:
    clean = (value or "").strip()
    return PRIORITY_LABELS.get(clean, clean.title() if clean else "")


def _slug_from_url(url: str) -> str:
    path = unquote(urlparse(url).path).rstrip("/")
    return path.split("/")[-1] if path else ""


def _normalize_url(url: str) -> str:
    parsed = urlparse(url)
    path = unquote(parsed.path).rstrip("/").lower()
    return f"{parsed.scheme}://{parsed.netloc.lower()}{path}"


def _sheet_name(category: str, used: Dict[str, int]) -> str:
    """Excel sheet names are limited to 31 chars and must be unique."""
    base = re.sub(r'[\[\]\*\?:/\\]', "", category).strip() or "Uncategorized"
    base = base[:31]
    if base not in used:
        used[base] = 0
        return base
    used[base] += 1
    suffix = f" ({used[base]})"
    return f"{base[: 31 - len(suffix)]}{suffix}"


def parse_post_sitemap(xml_text: str) -> List[SitemapEntry]:
    root = ET.fromstring(xml_text)
    entries: List[SitemapEntry] = []

    for node in root.findall("sm:url", SITEMAP_NS):
        loc_el = node.find("sm:loc", SITEMAP_NS)
        if loc_el is None or not loc_el.text:
            continue

        url = _strip_cdata(loc_el.text)
        lastmod_el = node.find("sm:lastmod", SITEMAP_NS)
        changefreq_el = node.find("sm:changefreq", SITEMAP_NS)
        priority_el = node.find("sm:priority", SITEMAP_NS)

        lastmod = _strip_cdata(lastmod_el.text if lastmod_el is not None else "")
        changefreq = _strip_cdata(changefreq_el.text if changefreq_el is not None else "")
        priority = _strip_cdata(priority_el.text if priority_el is not None else "")
        combined, date_part, time_part = _format_lastmod(lastmod)
        image_count = len(node.findall("img:image", SITEMAP_NS))

        entries.append(
            SitemapEntry(
                url=url,
                images=image_count,
                change_frequency=changefreq.title() if changefreq else "",
                priority=priority,
                priority_label=_priority_label(priority),
                last_updated=combined,
                last_updated_date=date_part,
                last_updated_time=time_part,
                slug=_slug_from_url(url),
            )
        )

    return entries


def _category_names(post: dict) -> Tuple[str, List[str]]:
    """Return primary category and any subcategories from a WP post payload."""
    categories: List[str] = []
    subcategories: List[str] = []

    for group in post.get("_embedded", {}).get("wp:term", []):
        for term in group:
            if term.get("taxonomy") != "category":
                continue
            name = (term.get("name") or "").strip()
            if not name:
                continue
            parent = term.get("parent", 0)
            if parent:
                subcategories.append(name)
            else:
                categories.append(name)

    primary = categories[0] if categories else (subcategories[0] if subcategories else "Uncategorized")
    if primary in subcategories and categories:
        subcategories = [s for s in subcategories if s != primary]

    # Preserve order, drop duplicates.
    seen = set()
    deduped_subs: List[str] = []
    for sub in subcategories:
        if sub not in seen and sub != primary:
            seen.add(sub)
            deduped_subs.append(sub)

    return primary, deduped_subs


def fetch_wp_posts(session: requests.Session) -> Dict[str, dict]:
    """Fetch all published posts from the WordPress REST API."""
    by_link: Dict[str, dict] = {}
    page = 1

    while True:
        response = session.get(
            WP_POSTS_URL,
            params={"per_page": 100, "page": page, "_embed": ""},
            timeout=30,
        )
        if response.status_code == 400:
            break
        response.raise_for_status()
        posts = response.json()
        if not posts:
            break

        for post in posts:
            link = (post.get("link") or "").strip()
            if not link:
                continue
            by_link[link] = post
            by_link[link.rstrip("/")] = post
            by_link[f"{link.rstrip('/')}/"] = post
            by_link[_normalize_url(link)] = post

        total_pages = int(response.headers.get("X-WP-TotalPages", page))
        if page >= total_pages:
            break
        page += 1

    return by_link


def enrich_entries(entries: List[SitemapEntry], wp_posts: Dict[str, dict]) -> None:
    by_slug = {
        (post.get("slug") or "").strip(): post
        for post in wp_posts.values()
        if post.get("slug")
    }

    for entry in entries:
        normalized_url = _normalize_url(entry.url)
        post = (
            wp_posts.get(entry.url)
            or wp_posts.get(entry.url.rstrip("/"))
            or wp_posts.get(f"{entry.url.rstrip('/')}/")
            or wp_posts.get(normalized_url)
            or by_slug.get(entry.slug)
        )
        if not post:
            continue

        title = post.get("title", {}).get("rendered", "")
        entry.title = re.sub(r"\s+", " ", html.unescape(title or "")).strip()
        entry.slug = post.get("slug") or entry.slug

        published = post.get("date", "")
        if published:
            try:
                dt = datetime.fromisoformat(published)
                entry.published_date = dt.strftime("%Y-%m-%d")
            except ValueError:
                entry.published_date = published[:10]

        entry.topic_category, entry.subcategories = _category_names(post)


def sort_entries_by_category(entries: List[SitemapEntry]) -> List[SitemapEntry]:
    """Group rows by topic category; largest clusters first, then title A-Z."""
    counts: Dict[str, int] = {}
    for entry in entries:
        counts[entry.topic_category] = counts.get(entry.topic_category, 0) + 1

    return sorted(
        entries,
        key=lambda entry: (
            -counts.get(entry.topic_category, 0),
            entry.topic_category.lower(),
            entry.title.lower(),
        ),
    )


def write_excel(entries: List[SitemapEntry], path: str) -> None:
    headers = [
        "Blog Title",
        "URL",
        "Slug",
        "Topic Category",
        "Subcategories",
        "Images",
        "Change Frequency",
        "Priority",
        "Priority Label",
        "Last Updated",
        "Last Updated Date",
        "Last Updated Time",
        "Published Date",
    ]

    wb = Workbook()
    used_sheet_names: Dict[str, int] = {}

    def style_header(ws, row: int = 1) -> None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def append_rows(ws, rows: List[SitemapEntry], start_row: int = 2) -> None:
        wrap = Alignment(wrap_text=True, vertical="top")
        link_font = Font(color="0563C1", underline="single")
        row_idx = start_row

        for entry in rows:
            ws.cell(row=row_idx, column=1, value=entry.title).alignment = wrap

            url_cell = ws.cell(row=row_idx, column=2, value=entry.url)
            url_cell.hyperlink = entry.url
            url_cell.font = link_font
            url_cell.alignment = wrap

            ws.cell(row=row_idx, column=3, value=entry.slug).alignment = wrap
            ws.cell(row=row_idx, column=4, value=entry.topic_category).alignment = wrap
            ws.cell(row=row_idx, column=5, value=", ".join(entry.subcategories)).alignment = wrap
            ws.cell(row=row_idx, column=6, value=entry.images).alignment = wrap
            ws.cell(row=row_idx, column=7, value=entry.change_frequency).alignment = wrap
            ws.cell(row=row_idx, column=8, value=entry.priority).alignment = wrap
            ws.cell(row=row_idx, column=9, value=entry.priority_label).alignment = wrap
            ws.cell(row=row_idx, column=10, value=entry.last_updated).alignment = wrap
            ws.cell(row=row_idx, column=11, value=entry.last_updated_date).alignment = wrap
            ws.cell(row=row_idx, column=12, value=entry.last_updated_time).alignment = wrap
            ws.cell(row=row_idx, column=13, value=entry.published_date).alignment = wrap
            row_idx += 1

    def autosize(ws) -> None:
        widths = {
            "Blog Title": 55,
            "URL": 65,
            "Slug": 35,
            "Topic Category": 22,
            "Subcategories": 28,
            "Images": 10,
            "Change Frequency": 16,
            "Priority": 10,
            "Priority Label": 14,
            "Last Updated": 24,
            "Last Updated Date": 18,
            "Last Updated Time": 14,
            "Published Date": 14,
        }
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 18)

    grouped_entries = sort_entries_by_category(entries)

    # Main sheet — all posts grouped by topic category
    ws_all = wb.active
    ws_all.title = "All Posts"
    ws_all.append(headers)
    style_header(ws_all)
    append_rows(ws_all, grouped_entries)
    autosize(ws_all)
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = ws_all.dimensions

    # Summary sheet
    ws_summary = wb.create_sheet("Category Summary")
    ws_summary.append(["Topic Category", "Post Count", "Share %"])
    style_header(ws_summary, row=1)
    ws_summary.cell(1, 1).value = "Topic Category"
    ws_summary.cell(1, 2).value = "Post Count"
    ws_summary.cell(1, 3).value = "Share %"

    counts: Dict[str, int] = {}
    for entry in entries:
        counts[entry.topic_category] = counts.get(entry.topic_category, 0) + 1

    total = len(entries) or 1
    summary_rows = sorted(counts.items(), key=lambda item: (-item[1], item[0].lower()))
    for row_idx, (category, count) in enumerate(summary_rows, start=2):
        ws_summary.cell(row=row_idx, column=1, value=category)
        ws_summary.cell(row=row_idx, column=2, value=count)
        ws_summary.cell(row=row_idx, column=3, value=round(100 * count / total, 1))
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.freeze_panes = "A2"

    # One sheet per topic cluster
    clusters: Dict[str, List[SitemapEntry]] = {}
    for entry in entries:
        clusters.setdefault(entry.topic_category, []).append(entry)

    for category in sorted(clusters, key=str.lower):
        sheet_name = _sheet_name(category, used_sheet_names)
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        style_header(ws)
        category_rows = sorted(clusters[category], key=lambda entry: entry.title.lower())
        append_rows(ws, category_rows)
        autosize(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Scrape ASBL post sitemap, cluster by topic category, export Excel."
    )
    parser.add_argument(
        "--sitemap-url",
        default=SITEMAP_URL,
        help=f"Sitemap XML URL. Default: {SITEMAP_URL}",
    )
    parser.add_argument(
        "--out",
        default="asbl_post_sitemap_clusters.xlsx",
        help="Output Excel workbook path.",
    )
    parser.add_argument(
        "--local-xml",
        default="",
        help="Optional local post-sitemap.xml file (skips live sitemap fetch).",
    )
    args = parser.parse_args()

    session = make_session()

    if args.local_xml:
        with open(args.local_xml, encoding="utf-8") as handle:
            xml_text = handle.read()
        if not xml_text.lstrip().startswith("<?xml"):
            start = xml_text.find("<?xml")
            xml_text = xml_text[start:] if start >= 0 else xml_text
    else:
        print(f"Fetching sitemap: {args.sitemap_url}", file=sys.stderr)
        xml_text = fetch_text(session, args.sitemap_url)

    entries = parse_post_sitemap(xml_text)
    print(f"Parsed {len(entries)} sitemap URLs.", file=sys.stderr)

    print("Fetching WordPress post metadata for topic categories...", file=sys.stderr)
    wp_posts = fetch_wp_posts(session)
    enrich_entries(entries, wp_posts)

    matched = sum(1 for entry in entries if entry.title)
    print(f"Matched {matched}/{len(entries)} posts to WordPress metadata.", file=sys.stderr)

    write_excel(entries, args.out)
    clusters = len({entry.topic_category for entry in entries})
    print(
        f"Saved {len(entries)} posts across {clusters} topic clusters -> {args.out}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
