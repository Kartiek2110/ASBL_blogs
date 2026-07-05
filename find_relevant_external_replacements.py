"""
Find verified, topically relevant external links to replace irrelevant ones.

Quality gates (every suggested replacement must pass ALL of these):
  1. URL is live — page loads and yields a real title (not 404/dead)
  2. Was marked "Relevant" somewhere in the ASBL blog audit, OR passes
     strict keyword relevance to the host blog after live title fetch
  3. Relevance score to the host blog >= --threshold (default 0.15)
  4. At least --min-keywords shared topic words (default 2)
  5. Not already present on the host blog

If no candidate passes all gates, the row is flagged NEEDS MANUAL REVIEW
rather than forcing a bad suggestion.

Input:  irrelevant_links.xlsx
Output: relevant_external_replacements.xlsx

Run:
    python find_relevant_external_replacements.py --workers 10
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm

from analyze_irrelevant_links import (
    DOMAIN_STOP,
    GENERIC_LABELS,
    _blog_keywords,
    _extract_page_title,
    _link_keywords,
    _normalize_label,
    _tokenize,
    score_relevance,
)
from scrape_asbl_blogs import (
    BlogListing,
    _clean_text,
    collect_listings,
    fetch,
    make_session,
)

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# Weak words that must NOT count toward "specific" topical overlap.
GENERIC_OVERLAP: Set[str] = DOMAIN_STOP | {
    "make", "help", "details", "financial", "plan", "plans", "long", "down",
    "family", "let", "change", "impact", "means", "even", "first", "main",
    "role", "post", "crucial", "different", "them", "safe", "technology",
    "area", "essential", "improve", "loan", "one", "plays", "good", "high",
    "city", "activity", "region", "designed", "list", "table", "top", "localities",
    "communities", "expect", "example", "less", "especially", "people",
    "preferences", "choices", "housing", "advantages", "maintain", "management",
    "robust", "residential", "game", "environmental", "development", "industrial",
    "letter", "credit", "modern", "floor", "ground", "water", "clean", "solutions",
    "treatment", "plants", "concrete", "resilience", "cement", "ways", "ten",
    "score", "cibil", "gated", "neighbourhood", "neighborhood", "complete",
    "avoid", "tips", "right", "better", "ultimate", "minute", "smart", "read",
    "full", "know", "what", "check", "move", "ready", "things", "keep", "paying",
    "money", "today", "tomorrow", "needs", "like", "space", "style", "discover",
    "explore", "understanding", "while", "important", "choose", "maximum",
    "ahead", "insights", "market", "deal", "growing", "start", "they", "year",
    "built", "practical", "spaces", "buyer", "benefits", "options", "valuable",
    "government", "single", "women", "mortgage", "vehicle", "rates", "interest",
    "small", "solution", "bill", "parties", "concept", "joint", "beyond", "coming",
    "metros", "projects", "whether", "luxury", "priced", "properties", "quickly",
    "planning", "lending", "finances", "calculate", "dive", "efficient", "maximize",
}

# Words that signal a foreign / off-region page when the blog is India-focused.
FOREIGN_LOCATIONS: Set[str] = {
    "houston", "texas", "california", "florida", "australia", "sydney", "melbourne",
    "canada", "toronto", "vancouver", "uk", "london", "manchester", "dubai", "uae",
    "singapore", "new york", "chicago", "boston", "seattle", "atlanta",
}

INDIA_LOCATIONS: Set[str] = {
    "hyderabad", "telangana", "india", "indian", "bangalore", "bengaluru", "mumbai",
    "delhi", "chennai", "kolkata", "pune", "gurgaon", "noida",
}

# Page-title phrases that are almost never relevant for ASBL real-estate blogs.
OFF_TOPIC_TITLE_PATTERNS: Tuple[str, ...] = (
    "interior paint",
    "exotic pet",
    "share price",
    "stock price",
    "2bhk interior",
    "3bhk floor plan",
    "airslate sign",
    "lending rates",
    "minute read",
    "designing today for a better tomorrow",
)

# Domains that consistently fail or are low-quality — never suggest these.
BLOCKED_DOMAINS: Set[str] = {
    "accio.com",
    "urbanriseopulence.org.in",
    "proprety.in",
    "vrogue.co",
    "biglive.com",
    "h-o.engineering",
    "hyderabadonline.in",
    "assethub.co.in",
    "leaddeveloper.com",
    "theurbanengine.com",
    "sakshipost.com",
    "lawyersonia.com",
    "iaai.in",
    "karma.law",
    "realtydatahub.com",
    "axley.com",
}


@dataclass
class BlogTopic:
    url: str
    title: str
    keywords: Set[str] = field(default_factory=set)
    title_keywords: Set[str] = field(default_factory=set)


@dataclass
class ValidatedExternal:
    url: str
    norm: str
    domain: str
    page_title: str
    http_status: int
    keywords: Set[str] = field(default_factory=set)
    audit_relevant: bool = False


@dataclass
class ReplacementRow:
    blog_name: str
    blog_link: str
    irrelevant_link_url: str
    irrelevant_link_text: str
    replacement_url: str
    replacement_title: str
    replacement_domain: str
    relevance_score: float
    shared_keywords: str
    relevance_reason: str
    url_status: str
    verified: str
    match_tier: str = ""


def _normalize_url(url: str) -> str:
    parsed = urlparse(str(url).strip())
    query = parse_qs(parsed.query)
    for key in ("utm_source", "utm_medium", "utm_campaign", "utm"):
        query.pop(key, None)
    clean_query = urlencode({k: v[0] for k, v in query.items()}, doseq=False)
    return urlunparse((
        parsed.scheme.lower(),
        parsed.netloc.lower(),
        parsed.path.rstrip("/"),
        parsed.params,
        clean_query,
        "",
    ))


def _domain(url: str) -> str:
    host = urlparse(url).netloc.lower()
    return host[4:] if host.startswith("www.") else host


def _is_blocked(url: str) -> bool:
    d = _domain(url)
    return d in BLOCKED_DOMAINS or any(d.endswith("." + b) for b in BLOCKED_DOMAINS)


def _extract_body_snippet(html: str, limit: int = 1500) -> str:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        return _clean_text(meta["content"])[:limit]
    article = soup.find("article") or soup.find("main") or soup.body
    if article:
        return _clean_text(article.get_text())[:limit]
    return ""


def _extract_blog_body(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    el = soup.select_one(".entry-content, .post_content")
    return _clean_text(el.get_text()) if el else ""


def _extract_blog_title(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    meta = soup.find("meta", property="og:title")
    if meta and meta.get("content"):
        return _clean_text(meta["content"])
    h1 = soup.find("h1", class_="entry-title") or soup.find("h1")
    return _clean_text(h1.get_text()) if h1 else ""


def validate_external_url(url: str) -> Tuple[bool, int, str, str]:
    """
    Fetch the page. Returns (is_valid, status_code, page_title, body_snippet).
    Valid = not 404, has extractable title, domain not blocked.
    """
    if _is_blocked(url):
        return False, 0, "", ""

    session = requests.Session()
    session.headers.update(BROWSER_HEADERS)

    try:
        resp = session.get(url, timeout=20, allow_redirects=True)
    except requests.RequestException:
        return False, 0, "", ""

    if resp.status_code != 200:
        return False, resp.status_code, "", ""

    html = resp.text or ""
    if len(html) < 200:
        return False, resp.status_code, "", ""

    title = _extract_page_title(html)
    snippet = _extract_body_snippet(html)

    if not title or len(title) < 8:
        return False, resp.status_code, "", snippet

    # Reject obvious error/placeholder pages
    lower_title = title.lower()
    if any(x in lower_title for x in ("404", "not found", "page not found", "access denied", "error")):
        return False, resp.status_code, title, snippet

    return True, resp.status_code, title, snippet


def load_cache(path: Path) -> Dict[str, dict]:
    if path.exists():
        try:
            return json.loads(path.read_text())
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_cache(path: Path, cache: Dict[str, dict]) -> None:
    path.write_text(json.dumps(cache, indent=2))


def validate_catalog(
    urls: Dict[str, dict],
    cache_path: Path,
    workers: int,
    cache_only: bool = False,
) -> List[ValidatedExternal]:
    """Validate every candidate URL; use disk cache to avoid re-fetching."""
    cache = load_cache(cache_path)
    validated: List[ValidatedExternal] = []

    to_fetch: List[Tuple[str, dict]] = []
    for norm, meta in urls.items():
        if norm in cache:
            c = cache[norm]
            if c.get("valid") and c.get("status") == 200:
                validated.append(ValidatedExternal(
                    url=meta["url"],
                    norm=norm,
                    domain=_domain(meta["url"]),
                    page_title=c["title"],
                    http_status=c.get("status", 200),
                    keywords=set(c.get("keywords", [])),
                    audit_relevant=meta.get("audit_relevant", False),
                ))
        elif not cache_only:
            to_fetch.append((norm, meta))

    print(f"  Cache hits: {len(validated)}, to validate: {len(to_fetch)}", file=sys.stderr)

    if cache_only or not to_fetch:
        return validated

    def _validate_one(item: Tuple[str, dict]) -> Tuple[str, dict]:
        norm, meta = item
        ok, status, title, snippet = validate_external_url(meta["url"])
        keywords: Set[str] = set()
        if ok:
            link_kw = _link_keywords(
                " ".join(meta.get("labels", [])),
                meta["url"],
                title + " " + snippet,
            )
            keywords = link_kw
        return norm, {
            "valid": ok,
            "status": status,
            "title": title,
            "snippet": snippet[:500],
            "keywords": sorted(keywords),
            "url": meta["url"],
            "audit_relevant": meta.get("audit_relevant", False),
        }

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_validate_one, item): item[0] for item in to_fetch}
        for fut in tqdm(as_completed(futures), total=len(futures), desc="Validating URLs"):
            norm, result = fut.result()
            cache[norm] = result
            if result["valid"]:
                validated.append(ValidatedExternal(
                    url=result["url"],
                    norm=norm,
                    domain=_domain(result["url"]),
                    page_title=result["title"],
                    http_status=result["status"],
                    keywords=set(result["keywords"]),
                    audit_relevant=result["audit_relevant"],
                ))

    save_cache(cache_path, cache)
    return validated


def scrape_blog_topic(session, listing: BlogListing) -> Optional[BlogTopic]:
    html = fetch(session, listing.url)
    if not html:
        return None
    title = _extract_blog_title(html) or listing.title or ""
    body = _extract_blog_body(html)
    keywords = _blog_keywords(title, body, listing.url)
    title_keywords = _title_topic_keywords(title)
    return BlogTopic(url=listing.url, title=title, keywords=keywords, title_keywords=title_keywords)


def _specific_overlap(blog: BlogTopic, entry: ValidatedExternal) -> Tuple[Set[str], Set[str]]:
    overlap = blog.keywords & entry.keywords
    specific = {w for w in overlap if w not in GENERIC_OVERLAP and len(w) > 3}
    return specific, overlap


def _title_topic_keywords(title: str) -> Set[str]:
    return {w for w in _tokenize(title) if w not in GENERIC_OVERLAP and len(w) > 2}


def _geographic_mismatch(blog: BlogTopic, entry: ValidatedExternal) -> bool:
    blog_locs = _title_topic_keywords(blog.title) & INDIA_LOCATIONS
    if not blog_locs:
        return False
    link_text = (entry.page_title + " " + entry.url).lower()
    link_tokens = set(_tokenize(link_text))
    if link_tokens & FOREIGN_LOCATIONS:
        if not (link_tokens & INDIA_LOCATIONS):
            return True
    return False


def _off_topic_title(blog: BlogTopic, entry: ValidatedExternal) -> bool:
    lower = entry.page_title.lower()
    for pattern in OFF_TOPIC_TITLE_PATTERNS:
        if pattern in lower:
            return True
    # Stock/market ticker pages rarely fit editorial blog context.
    if "share price" in lower or "stock" in lower.split():
        blog_wants_market = {"stock", "share", "invest", "portfolio", "market"} & blog.title_keywords
        if not blog_wants_market:
            return True
    return False


def _anchor_keywords(label: str, url: str) -> Set[str]:
    clean = _normalize_label(label)
    if not clean or clean in GENERIC_LABELS:
        return set()
    return _link_keywords(label, url, "")


def _strict_match(
    blog: BlogTopic,
    entry: ValidatedExternal,
    threshold: float,
    min_specific: int,
    anchor_kw: Optional[Set[str]] = None,
    min_verified_score: float = 0.38,
    min_title_overlap: int = 2,
    allow_single_strong_title: bool = True,
) -> Optional[Tuple[float, str, Set[str]]]:
    """Return (score, reason, specific_keywords) or None if not a good match."""
    if _geographic_mismatch(blog, entry):
        return None
    if _off_topic_title(blog, entry):
        return None

    score, reason = score_relevance(blog.keywords, entry.keywords)
    specific, _ = _specific_overlap(blog, entry)

    if score < threshold:
        return None
    if len(specific) < min_specific:
        return None

    blog_title_kw = _title_topic_keywords(blog.title)
    link_title_kw = _title_topic_keywords(entry.page_title)
    title_overlap = blog_title_kw & link_title_kw
    strong_title_overlap = {w for w in title_overlap if len(w) >= 5}
    if len(title_overlap) < min_title_overlap:
        if not (allow_single_strong_title and strong_title_overlap):
            return None

    # Must share at least one specific keyword with the linked page title
    if not (specific & link_title_kw):
        return None

    # Only enforce anchor alignment when the irrelevant link text is clearly topical.
    if anchor_kw:
        anchor_meaningful = {
            w for w in anchor_kw if w not in GENERIC_OVERLAP and len(w) > 3
        }
        if len(anchor_meaningful) >= 2:
            anchor_score, _ = score_relevance(anchor_kw, entry.keywords)
            anchor_overlap = anchor_kw & entry.keywords
            anchor_specific = {
                w for w in anchor_overlap if w not in GENERIC_OVERLAP and len(w) > 3
            }
            if anchor_score < 0.10 and len(anchor_specific) < 1:
                return None

    specific_score = len(specific) / min(len(blog.keywords), len(entry.keywords), 12)
    title_bonus = min(len(title_overlap) * 0.08, 0.25)
    combined = min(specific_score + title_bonus, 1.0)
    if combined < min_verified_score:
        return None

    return combined, reason, specific


class BlogTopicMatcher:
    """Find blogs with similar topics."""

    def __init__(self, blogs: List[BlogTopic]):
        self.by_url = {b.url.rstrip("/").lower(): b for b in blogs}
        self.blogs = blogs

    def similarity(self, host: BlogTopic, other: BlogTopic) -> float:
        if not host.keywords or not other.keywords:
            return 0.0
        overlap = host.keywords & other.keywords
        specific = {w for w in overlap if w not in GENERIC_OVERLAP}
        if len(specific) < 2:
            return 0.0
        return len(specific) / min(len(host.keywords), len(other.keywords), 15)

    def similar_blogs(self, host: BlogTopic, top_n: int = 40) -> List[Tuple[BlogTopic, float]]:
        host_key = host.url.rstrip("/").lower()
        scored: List[Tuple[BlogTopic, float]] = []
        for blog in self.blogs:
            if blog.url.rstrip("/").lower() == host_key:
                continue
            sim = self.similarity(host, blog)
            if sim >= 0.06:
                scored.append((blog, sim))
        scored.sort(key=lambda x: -x[1])
        return scored[:top_n]


def build_url_sources(all_links_df: pd.DataFrame) -> Dict[str, Set[str]]:
    """Map each external URL to blog URLs where it was marked Relevant."""
    sources: Dict[str, Set[str]] = defaultdict(set)
    relevant = all_links_df[
        (all_links_df["Link Type"] == "External") & (all_links_df["Verdict"] == "Relevant")
    ]
    for _, row in relevant.iterrows():
        raw = str(row["Link URL"]).strip()
        if not raw.startswith("http"):
            continue
        blog = str(row["Blog Link"]).strip()
        sources[_normalize_url(raw)].add(blog.rstrip("/").lower())
    return sources


def find_replacements_for_blog(
    blog: BlogTopic,
    num_needed: int,
    exclude_norms: Set[str],
    catalog_by_norm: Dict[str, ValidatedExternal],
    url_sources: Dict[str, Set[str]],
    blog_matcher: BlogTopicMatcher,
    threshold: float,
    min_specific: int,
    global_usage: Counter,
    max_usage: int,
    anchor_kw: Optional[Set[str]] = None,
    min_verified_score: float = 0.38,
    min_title_overlap: int = 2,
) -> List[Tuple[ValidatedExternal, float, str, Set[str]]]:
    """Find validated, strictly relevant external links for one blog."""
    similar = blog_matcher.similar_blogs(blog, top_n=40)
    similar_keys = {b.url.rstrip("/").lower(): sim for b, sim in similar}

    # Priority 1: links already proven relevant on a similar-topic blog
    priority_norms: Set[str] = set()
    for norm, source_blogs in url_sources.items():
        if norm in exclude_norms or norm not in catalog_by_norm:
            continue
        for src in source_blogs:
            if src in similar_keys:
                priority_norms.add(norm)
                break

    def _score_candidates(norms: Set[str]) -> List[Tuple[ValidatedExternal, float, str, Set[str]]]:
        candidates: List[Tuple[ValidatedExternal, float, str, Set[str], float]] = []
        for norm in norms:
            entry = catalog_by_norm.get(norm)
            if not entry:
                continue
            matched = _strict_match(
                blog, entry, threshold, min_specific, anchor_kw, min_verified_score,
                min_title_overlap=min_title_overlap,
            )
            if not matched:
                continue
            spec_score, reason, specific = matched
            # Boost if link was relevant on a highly similar blog
            src_sims = [
                similar_keys[s] for s in url_sources.get(norm, set()) if s in similar_keys
            ]
            boost = max(src_sims) if src_sims else 0.0
            combined = min(spec_score * (1.0 + boost), 1.0)
            if combined < min_verified_score:
                continue
            candidates.append((entry, combined, reason, specific, boost))
        candidates.sort(key=lambda x: (-x[1], -len(x[3])))
        return [(e, s, r, k) for e, s, r, k, _ in candidates]

    scored = _score_candidates(priority_norms)

    # Priority 2: fall back to full validated catalog
    if len(scored) < num_needed:
        remaining = set(catalog_by_norm.keys()) - exclude_norms - priority_norms
        scored.extend(_score_candidates(remaining))

    results: List[Tuple[ValidatedExternal, float, str, Set[str]]] = []
    used: Set[str] = set()

    for entry, score, reason, specific in scored:
        if len(results) >= num_needed:
            break
        if entry.norm in used:
            continue
        if global_usage[entry.norm] >= max_usage:
            continue
        results.append((entry, score, reason, specific))
        used.add(entry.norm)
        global_usage[entry.norm] += 1

    return results


def build_candidate_urls(all_links_df: pd.DataFrame) -> Dict[str, dict]:
    """Collect unique external URLs from the audit (any verdict)."""
    by_url: Dict[str, dict] = {}

    external = all_links_df[all_links_df["Link Type"] == "External"]
    for _, row in external.iterrows():
        raw = str(row["Link URL"]).strip()
        if not raw.startswith("http"):
            continue
        if _is_blocked(raw):
            continue
        norm = _normalize_url(raw)
        label = "" if pd.isna(row["Link Text"]) else str(row["Link Text"]).strip()
        verdict = str(row.get("Verdict", "")).strip()
        if norm not in by_url:
            by_url[norm] = {
                "url": raw,
                "labels": [],
                "audit_relevant": verdict == "Relevant",
            }
        else:
            if verdict == "Relevant":
                by_url[norm]["audit_relevant"] = True
        if label and label not in by_url[norm]["labels"]:
            by_url[norm]["labels"].append(label)

    return by_url


def load_irrelevant_external(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Irrelevant Links")
    return df[df["Link Type"] == "External"].copy()


def confirm_url_live(url: str) -> bool:
    if not url.startswith("http") or _is_blocked(url):
        return False
    try:
        resp = requests.head(url, headers=BROWSER_HEADERS, timeout=15, allow_redirects=True)
        if resp.status_code >= 400:
            resp = requests.get(url, headers=BROWSER_HEADERS, timeout=15, allow_redirects=True, stream=True)
        return resp.status_code == 200
    except requests.RequestException:
        return False


def write_excel(rows: List[ReplacementRow], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "External Replacements"

    headers = [
        "Blog Name",
        "Blog Link",
        "Irrelevant Link Text",
        "Irrelevant Link URL",
        "Suggested Replacement URL",
        "Suggested Replacement Title",
        "Replacement Domain",
        "Relevance Score",
        "Shared Keywords",
        "Relevance Reason",
        "URL Status",
        "Verified",
        "Match Tier",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1565C0")
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")
    good_fill = PatternFill("solid", fgColor="C8E6C9")
    irr_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFF9C4")

    for row in rows:
        ri = ws.max_row + 1
        ws.cell(row=ri, column=1, value=row.blog_name).alignment = wrap

        bc = ws.cell(row=ri, column=2, value=row.blog_link)
        bc.hyperlink = row.blog_link
        bc.font = link_font
        bc.alignment = wrap

        ws.cell(row=ri, column=3, value=row.irrelevant_link_text).alignment = wrap

        ic = ws.cell(row=ri, column=4, value=row.irrelevant_link_url)
        ic.hyperlink = row.irrelevant_link_url
        ic.font = link_font
        ic.alignment = wrap
        ic.fill = irr_fill

        rc = ws.cell(row=ri, column=5, value=row.replacement_url)
        if row.replacement_url.startswith("http"):
            rc.hyperlink = row.replacement_url
            rc.fill = good_fill
        else:
            for c in range(1, 14):
                ws.cell(row=ri, column=c).fill = warn_fill
        rc.font = link_font
        rc.alignment = wrap

        ws.cell(row=ri, column=6, value=row.replacement_title).alignment = wrap
        ws.cell(row=ri, column=7, value=row.replacement_domain).alignment = wrap
        ws.cell(row=ri, column=8, value=row.relevance_score).alignment = wrap
        ws.cell(row=ri, column=9, value=row.shared_keywords).alignment = wrap
        ws.cell(row=ri, column=10, value=row.relevance_reason).alignment = wrap
        ws.cell(row=ri, column=11, value=row.url_status).alignment = wrap
        ws.cell(row=ri, column=12, value=row.verified).alignment = wrap
        ws.cell(row=ri, column=13, value=row.match_tier).alignment = wrap

    widths = [45, 60, 35, 60, 60, 45, 28, 14, 35, 50, 12, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Find verified relevant external link replacements.")
    parser.add_argument("--input", default="irrelevant_links.xlsx")
    parser.add_argument("--out", default="relevant_external_replacements.xlsx")
    parser.add_argument("--start", type=int, default=1)
    parser.add_argument("--end", type=int, default=43)
    parser.add_argument("--workers", type=int, default=10)
    parser.add_argument("--threshold", type=float, default=0.25,
                        help="Min audit-style relevance score. Default: 0.25")
    parser.add_argument("--min-keywords", type=int, default=2,
                        help="Min specific (non-generic) shared keywords. Default: 2")
    parser.add_argument("--min-verified-score", type=float, default=0.38,
                        help="Min combined score to mark Verified Yes. Default: 0.38")
    parser.add_argument("--max-usage", type=int, default=10,
                        help="Max times one URL is reused globally. Default: 10")
    parser.add_argument("--cache", default="external_url_cache.json")
    parser.add_argument("--fast", action="store_true",
                        help="Use cached URLs only, skip blog scrape and live re-checks")
    args = parser.parse_args()

    cache_path = Path(args.cache)

    print("Step 1: Loading irrelevant external links...", file=sys.stderr)
    irr_df = load_irrelevant_external(args.input)
    print(f"  {len(irr_df)} links across {irr_df['Blog Link'].nunique()} blogs", file=sys.stderr)

    print("\nStep 2: Building candidate URL list (all external URLs from audit)...", file=sys.stderr)
    all_links_df = pd.read_excel(args.input, sheet_name="All Links")
    candidates = build_candidate_urls(all_links_df)
    print(f"  {len(candidates)} unique external URLs in audit", file=sys.stderr)

    print("\nStep 3: Validating all candidate URLs (live check + title fetch)...", file=sys.stderr)
    validated_catalog = validate_catalog(
        candidates, cache_path, args.workers, cache_only=args.fast,
    )
    validated_catalog = [e for e in validated_catalog if not _is_blocked(e.url)]
    print(f"  {len(validated_catalog)} URLs passed validation (live + title)", file=sys.stderr)

    if not validated_catalog:
        print("ERROR: No valid external URLs in catalog. Aborting.", file=sys.stderr)
        return 1

    blog_topics: Dict[str, BlogTopic] = {}

    if args.fast:
        print("\nStep 4: Using blog titles from audit (fast mode)...", file=sys.stderr)
        for blog_link, group in irr_df.groupby("Blog Link"):
            blog_name = group.iloc[0]["Blog Name"]
            key = str(blog_link).rstrip("/").lower()
            blog_topics[key] = BlogTopic(
                url=str(blog_link),
                title=blog_name,
                keywords=_blog_keywords(blog_name, "", str(blog_link)),
                title_keywords=_title_topic_keywords(blog_name),
            )
        print(f"  {len(blog_topics)} blogs indexed from audit titles", file=sys.stderr)
    else:
        print("\nStep 4: Scraping ASBL blogs for topic keywords...", file=sys.stderr)
        session = make_session()
        listings = collect_listings(session, args.start, args.end)

        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures = {pool.submit(scrape_blog_topic, make_session(), e): e for e in listings}
            for fut in tqdm(as_completed(futures), total=len(futures), desc="Blog topics"):
                try:
                    topic = fut.result()
                except Exception:
                    continue
                if topic:
                    blog_topics[topic.url.rstrip("/").lower()] = topic

        print(f"  {len(blog_topics)} blogs indexed", file=sys.stderr)

    catalog_by_norm = {e.norm: e for e in validated_catalog}
    url_sources = build_url_sources(all_links_df)
    blog_matcher = BlogTopicMatcher(list(blog_topics.values()))

    print("\nStep 5: Matching replacements (strict relevance)...", file=sys.stderr)
    rows: List[ReplacementRow] = []
    global_usage: Counter = Counter()
    used_per_blog: Dict[str, Set[str]] = defaultdict(set)

    # Precompute existing external links per blog (to exclude)
    blog_existing: Dict[str, Set[str]] = defaultdict(set)
    for _, link_row in all_links_df[all_links_df["Link Type"] == "External"].iterrows():
        blog_existing[str(link_row["Blog Link"]).rstrip("/").lower()].add(
            _normalize_url(link_row["Link URL"])
        )

    for _, irr_row in tqdm(irr_df.iterrows(), total=len(irr_df), desc="Matching"):
        blog_link = str(irr_row["Blog Link"]).strip()
        blog_name = irr_row["Blog Name"]
        blog_key = blog_link.rstrip("/").lower()
        blog = blog_topics.get(blog_key)

        if not blog:
            blog = BlogTopic(
                url=blog_link,
                title=blog_name,
                keywords=_blog_keywords(blog_name, "", blog_link),
                title_keywords=_title_topic_keywords(blog_name),
            )

        exclude = set(blog_existing.get(blog_key, set()))
        exclude.add(_normalize_url(irr_row["Link URL"]))
        exclude |= used_per_blog[blog_key]

        anchor_kw = _anchor_keywords(
            "" if pd.isna(irr_row["Link Text"]) else str(irr_row["Link Text"]),
            str(irr_row["Link URL"]),
        )

        match_tier = ""
        chosen = None

        pass_configs = [
            ("Strict", {
                "threshold": args.threshold,
                "min_specific": args.min_keywords,
                "min_verified_score": args.min_verified_score,
                "min_title_overlap": 2,
            }),
            ("Relaxed", {
                "threshold": 0.18,
                "min_specific": 2,
                "min_verified_score": 0.32,
                "min_title_overlap": 1,
            }),
        ]

        for tier_name, cfg in pass_configs:
            replacements = find_replacements_for_blog(
                blog=blog,
                num_needed=8,
                exclude_norms=exclude,
                catalog_by_norm=catalog_by_norm,
                url_sources=url_sources,
                blog_matcher=blog_matcher,
                threshold=cfg["threshold"],
                min_specific=cfg["min_specific"],
                global_usage=global_usage,
                max_usage=args.max_usage,
                anchor_kw=anchor_kw or None,
                min_verified_score=cfg["min_verified_score"],
                min_title_overlap=cfg["min_title_overlap"],
            )
            for entry, score, reason, overlap in replacements:
                if args.fast or confirm_url_live(entry.url):
                    chosen = (entry, score, reason, overlap)
                    match_tier = tier_name
                    used_per_blog[blog_key].add(entry.norm)
                    break
            if chosen:
                break

        if chosen:
            entry, score, reason, overlap = chosen
            rows.append(ReplacementRow(
                blog_name=blog_name,
                blog_link=blog_link,
                irrelevant_link_url=irr_row["Link URL"],
                irrelevant_link_text=irr_row["Link Text"],
                replacement_url=entry.url,
                replacement_title=entry.page_title,
                replacement_domain=entry.domain,
                relevance_score=round(score, 3),
                shared_keywords=", ".join(sorted(overlap)[:8]),
                relevance_reason=reason,
                url_status=str(entry.http_status),
                verified="Yes",
                match_tier=match_tier,
            ))
        else:
            rows.append(ReplacementRow(
                blog_name=blog_name,
                blog_link=blog_link,
                irrelevant_link_url=irr_row["Link URL"],
                irrelevant_link_text=irr_row["Link Text"],
                replacement_url="NEEDS MANUAL REVIEW",
                replacement_title="",
                replacement_domain="",
                relevance_score=0.0,
                shared_keywords="",
                relevance_reason="No live, relevant external URL found in validated pool",
                url_status="",
                verified="No",
                match_tier="",
            ))

    print(f"\nStep 6: Writing {args.out}...", file=sys.stderr)
    write_excel(rows, args.out)

    verified = sum(1 for r in rows if r.verified == "Yes")
    manual = len(rows) - verified
    unique_urls = len({r.replacement_url for r in rows if r.verified == "Yes"})

    print(
        f"\nDone!\n"
        f"  Total irrelevant external links:  {len(rows)}\n"
        f"  Verified replacements:            {verified}\n"
        f"  Needs manual review:                {manual}\n"
        f"  Unique replacement URLs used:       {unique_urls}\n"
        f"  Validated catalog size:             {len(validated_catalog)}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
