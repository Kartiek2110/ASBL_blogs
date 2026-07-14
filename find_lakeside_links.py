"""
Find all ASBL blog posts that link to (or mention) Lakeside in internal links.

Scans every blog listed in internal_links.xlsx plus the full sitemap catalog,
visits each post, extracts internal links from the article body, and records
every Lakeside-related link with where it was found.

Run:
    python find_lakeside_links.py
    python find_lakeside_links.py --in "/path/to/internal_links.xlsx" --out lakeside_links_report.xlsx
"""

from __future__ import annotations

import argparse
import html
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import List, Optional, Set, Tuple
from urllib.parse import unquote, urlparse

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from scrape_asbl_blogs import _clean_text, fetch, make_session

SEARCH_TERM = "lakeside"
INTERNAL_HOST = "asbl.in"

HEADERS = [
    "Source Blog Title",
    "Source Blog URL",
    "Source Blog Date",
    "Lakeside Link Text",
    "Lakeside Link URL",
    "Match In",
    "Link Type",
    "Notes",
]

COLUMN_WIDTHS = {
    "Source Blog Title": 52,
    "Source Blog URL": 62,
    "Source Blog Date": 14,
    "Lakeside Link Text": 40,
    "Lakeside Link URL": 62,
    "Match In": 22,
    "Link Type": 14,
    "Notes": 36,
}


@dataclass(frozen=True)
class ExcelStyles:
    header_font: Font
    header_fill: PatternFill
    wrap: Alignment
    link_font: Font


def _excel_styles() -> ExcelStyles:
    return ExcelStyles(
        header_font=Font(bold=True, color="FFFFFF"),
        header_fill=PatternFill("solid", fgColor="1F4E78"),
        wrap=Alignment(wrap_text=True, vertical="top"),
        link_font=Font(color="0563C1", underline="single"),
    )


@dataclass
class LakesideHit:
    source_title: str
    source_url: str
    source_date: str
    link_text: str
    link_url: str
    match_in: str
    link_type: str
    notes: str

    def as_row(self) -> List[str]:
        return [
            self.source_title,
            self.source_url,
            self.source_date,
            self.link_text,
            self.link_url,
            self.match_in,
            self.link_type,
            self.notes,
        ]

    @classmethod
    def lakeside_blog(cls, title: str, url: str, date: str) -> LakesideHit:
        return cls(
            source_title=title,
            source_url=url,
            source_date=date,
            link_text="(This blog is about Lakeside)",
            link_url=url,
            match_in="Blog title/URL",
            link_type="Lakeside blog",
            notes="Blog title or slug references Lakeside",
        )

    @classmethod
    def from_anchor(
        cls,
        title: str,
        page_url: str,
        date: str,
        text: str,
        href: str,
    ) -> LakesideHit:
        internal = _is_internal(href)
        return cls(
            source_title=title,
            source_url=page_url,
            source_date=date,
            link_text=text or href,
            link_url=href,
            match_in=_match_location(text, href),
            link_type="Internal" if internal else "External",
            notes="Found in article body",
        )

    @classmethod
    def text_mention(cls, title: str, url: str, date: str) -> LakesideHit:
        return cls(
            source_title=title,
            source_url=url,
            source_date=date,
            link_text="(Text mention only — no clickable Lakeside link)",
            link_url="",
            match_in="Article body text",
            link_type="Text mention",
            notes="Lakeside mentioned in blog body without a dedicated link",
        )


def _is_internal(url: str) -> bool:
    try:
        host = urlparse(url).netloc.lower()
    except ValueError:
        return False
    return host == INTERNAL_HOST or host.endswith("." + INTERNAL_HOST)


def _extract_title(soup: BeautifulSoup) -> str:
    h1 = soup.find("h1", class_="entry-title") or soup.find("h1")
    if h1:
        return _clean_text(h1.get_text())
    meta = soup.find("meta", property="og:title")
    if meta and meta.get("content"):
        return _clean_text(meta["content"])
    return ""


def _extract_date(soup: BeautifulSoup) -> str:
    meta = soup.find("meta", property="article:published_time")
    if meta and meta.get("content"):
        return meta["content"][:10]
    t = soup.find("time")
    if t and t.get("datetime"):
        return t["datetime"][:10]
    return ""


def _contains_lakeside(value: str) -> bool:
    return SEARCH_TERM in unquote(value or "").lower()


def _match_location(link_text: str, link_url: str) -> str:
    parts: List[str] = []
    if _contains_lakeside(link_url):
        parts.append("Link URL")
    if _contains_lakeside(link_text):
        parts.append("Link Text")
    return " + ".join(parts) if parts else "Page body"


def _article_content(soup: BeautifulSoup):
    return soup.select_one(".entry-content") or soup.select_one(".post_content")


def _add_self_lakeside_hit(
    hits: List[LakesideHit],
    seen: Set[Tuple[str, str]],
    title: str,
    page_url: str,
    date: str,
) -> None:
    if not (_contains_lakeside(title) or _contains_lakeside(page_url)):
        return

    key = ("__self__", page_url)
    if key in seen:
        return

    seen.add(key)
    hits.append(LakesideHit.lakeside_blog(title, page_url, date))


def _add_anchor_hits(
    hits: List[LakesideHit],
    seen: Set[Tuple[str, str]],
    content,
    title: str,
    page_url: str,
    date: str,
) -> None:
    for anchor in content.find_all("a", href=True):
        href = html.unescape(anchor["href"].strip())
        text = _clean_text(anchor.get_text(" ", strip=True))
        if not _contains_lakeside(href) and not _contains_lakeside(text):
            continue

        key = (text, href)
        if key in seen:
            continue
        seen.add(key)
        hits.append(LakesideHit.from_anchor(title, page_url, date, text, href))


def _add_text_mention_hit(
    hits: List[LakesideHit],
    content,
    title: str,
    page_url: str,
    date: str,
) -> None:
    body_text = _clean_text(content.get_text(" ", strip=True)).lower()
    has_non_blog_link = any(h.link_url for h in hits if h.link_type != "Lakeside blog")
    if not _contains_lakeside(body_text) or has_non_blog_link or hits:
        return

    hits.append(LakesideHit.text_mention(title, page_url, date))


def _scan_html(
    page_url: str,
    page_html: str,
    source_date: str = "",
) -> List[LakesideHit]:
    soup = BeautifulSoup(page_html, "lxml")
    title = _extract_title(soup) or page_url
    date = source_date or _extract_date(soup)

    content = _article_content(soup)
    if content is None:
        return []

    hits: List[LakesideHit] = []
    seen: Set[Tuple[str, str]] = set()

    _add_self_lakeside_hit(hits, seen, title, page_url, date)
    _add_anchor_hits(hits, seen, content, title, page_url, date)
    _add_text_mention_hit(hits, content, title, page_url, date)

    return hits


def scan_blog_url(session, url: str, source_date: str = "") -> Tuple[Optional[str], List[LakesideHit]]:
    page_html = fetch(session, url, retries=4, timeout=35)
    if not page_html:
        return None, []
    if SEARCH_TERM not in page_html.lower():
        return url, []
    return url, _scan_html(url, page_html, source_date=source_date)


def _rows_from_sheet(
    df: pd.DataFrame,
    url_col: str,
    date_col: str,
    source: str,
) -> List[dict]:
    rows = []
    for _, row in df.iterrows():
        blog_url = str(row.get(url_col, "") or "").strip()
        if blog_url.startswith("http"):
            rows.append(
                {
                    "url": blog_url,
                    "date": str(row.get(date_col, "") or "")[:10],
                    "from": source,
                }
            )
    return rows


def collect_urls(input_path: str, catalog_path: str) -> pd.DataFrame:
    """Return unique blog URLs with optional dates from input + full catalog."""
    rows = _rows_from_sheet(
        pd.read_excel(input_path),
        url_col="Blog Link",
        date_col="Blog Date",
        source="internal_links.xlsx",
    )

    try:
        rows.extend(
            _rows_from_sheet(
                pd.read_excel(catalog_path, sheet_name="All Posts"),
                url_col="URL",
                date_col="Published Date",
                source="sitemap catalog",
            )
        )
    except Exception:
        pass

    frame = pd.DataFrame(rows)
    return frame.drop_duplicates(subset=["url"], keep="first")


def _style_header_row(ws: Worksheet, col_count: int, styles: ExcelStyles) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = styles.header_font
        cell.fill = styles.header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_hyperlink(cell, styles: ExcelStyles) -> None:
    if cell.value and str(cell.value).startswith("http"):
        cell.hyperlink = str(cell.value)
        cell.font = styles.link_font


def _write_hits_sheet(ws: Worksheet, hits: List[LakesideHit], styles: ExcelStyles) -> Set[str]:
    ws.title = "Lakeside Links"
    ws.append(HEADERS)
    _style_header_row(ws, len(HEADERS), styles)

    unique_blogs = {h.source_url for h in hits}
    for hit in hits:
        ws.append(hit.as_row())
        row_idx = ws.max_row
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = styles.wrap

        url_cell = ws.cell(row=row_idx, column=2)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.font = styles.link_font

        _apply_hyperlink(ws.cell(row=row_idx, column=5), styles)

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 20)

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    return unique_blogs


def _write_summary_sheet(
    wb: Workbook,
    scanned: int,
    unique_blogs: Set[str],
    hit_count: int,
    failed: List[str],
    styles: ExcelStyles,
) -> None:
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    for col_idx in (1, 2):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = styles.header_font
        cell.fill = styles.header_fill

    summary = [
        ("Search term", SEARCH_TERM),
        ("Blogs scanned", scanned),
        ("Blogs with Lakeside links/mentions", len(unique_blogs)),
        ("Total Lakeside link rows", hit_count),
        ("Failed to fetch", len(failed)),
    ]
    for idx, (metric, value) in enumerate(summary, start=2):
        ws_summary.cell(row=idx, column=1, value=metric)
        ws_summary.cell(row=idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 34
    ws_summary.column_dimensions["B"].width = 20


def _write_failed_sheet(wb: Workbook, failed: List[str], styles: ExcelStyles) -> None:
    ws_failed = wb.create_sheet("Failed URLs")
    ws_failed.append(["URL"])
    ws_failed.cell(1, 1).font = styles.header_font
    ws_failed.cell(1, 1).fill = styles.header_fill
    for url in failed:
        ws_failed.append([url])
    ws_failed.column_dimensions["A"].width = 70


def write_report(hits: List[LakesideHit], scanned: int, failed: List[str], path: str) -> None:
    wb = Workbook()
    styles = _excel_styles()

    unique_blogs = _write_hits_sheet(wb.active, hits, styles)
    _write_summary_sheet(wb, scanned, unique_blogs, len(hits), failed, styles)

    if failed:
        _write_failed_sheet(wb, failed, styles)

    wb.save(path)


def _scan_job(row) -> Tuple[str, Optional[str], List[LakesideHit]]:
    session = make_session()
    url = row["url"]
    try:
        status, result = scan_blog_url(session, url, source_date=row.get("date", ""))
        return url, status, result
    except Exception:
        return url, None, []


def scan_all_blogs(url_frame: pd.DataFrame, workers: int) -> Tuple[List[LakesideHit], List[str]]:
    hits: List[LakesideHit] = []
    failed: List[str] = []

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(_scan_job, row) for _, row in url_frame.iterrows()]
        for fut in tqdm(as_completed(futures), total=len(futures), desc="Scanning blogs"):
            url, status, result = fut.result()
            if status is None:
                failed.append(url)
            if result:
                hits.extend(result)

    hits.sort(key=lambda h: (h.source_title.lower(), h.link_url.lower()))
    return hits, failed


def main() -> int:
    parser = argparse.ArgumentParser(description="Find blogs linking to Lakeside.")
    parser.add_argument(
        "--in",
        dest="input_path",
        default="/Users/kartiekeybhardwaj/Downloads/rewindDesgin/internal_links.xlsx",
    )
    parser.add_argument(
        "--catalog",
        default="asbl_post_sitemap_clusters.xlsx",
        help="Full blog catalog for comprehensive scan.",
    )
    parser.add_argument(
        "--out",
        default="lakeside_links_report.xlsx",
    )
    parser.add_argument("--workers", type=int, default=6)
    args = parser.parse_args()

    url_frame = collect_urls(args.input_path, args.catalog)
    print(f"Scanning {len(url_frame)} unique blog URLs for '{SEARCH_TERM}'...", file=sys.stderr)

    hits, failed = scan_all_blogs(url_frame, args.workers)
    write_report(hits, scanned=len(url_frame), failed=failed, path=args.out)

    unique_blogs = len({h.source_url for h in hits})
    print(
        f"Done. {unique_blogs} blogs with Lakeside links/mentions, "
        f"{len(hits)} total rows -> {args.out}",
        file=sys.stderr,
    )
    if failed:
        print(f"{len(failed)} URLs failed to fetch (see Failed URLs sheet).", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
