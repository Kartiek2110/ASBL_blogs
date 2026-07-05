"""
Scan ASBL blog posts, score each styled internal/external link for topical
relevance to the host article, and export an Excel report for review.

Sheets:
  - Irrelevant Links  : one row per link flagged as not relevant to the blog
  - Blog Summary      : per-blog counts; blogs with irrelevant links listed first
  - All Links         : full audit trail (optional detail for every link scored)

Run:
    python analyze_irrelevant_links.py
    python analyze_irrelevant_links.py --start 1 --end 43 --workers 8 --out irrelevant_links.xlsx
    python analyze_irrelevant_links.py --limit 20   # quick preview
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from tqdm import tqdm

from scrape_asbl_blogs import (
    BlogListing,
    BlogRecord,
    _clean_text,
    collect_listings,
    fetch,
    make_session,
    parse_blog_post,
    scrape_post,
)

STOP_WORDS: Set[str] = {
    "a", "an", "the", "and", "or", "but", "in", "on", "at", "to", "for", "of",
    "is", "are", "was", "were", "be", "been", "being", "have", "has", "had",
    "do", "does", "did", "will", "would", "could", "should", "may", "might",
    "must", "shall", "can", "need", "your", "you", "our", "their", "its", "it",
    "this", "that", "these", "those", "with", "from", "by", "as", "about", "into",
    "through", "during", "before", "after", "above", "below", "between", "under",
    "again", "further", "then", "once", "here", "there", "when", "where", "why",
    "how", "all", "each", "few", "more", "most", "other", "some", "such", "no",
    "nor", "not", "only", "own", "same", "so", "than", "too", "very", "just",
    "also", "now", "new", "get", "blog", "asbl", "https", "http", "www", "com",
    "source", "related", "guide", "article", "html", "utm", "openai",
}

GENERIC_LABELS: Set[str] = {
    "source", "related", "read more", "click here", "here", "link",
    "related analysis", "learn more", "related reading", "see also",
}

DOMAIN_STOP: Set[str] = {
    "hyderabad", "india", "telangana", "home", "house", "property", "real",
    "estate", "apartment", "flats", "flat", "buy", "buying", "living", "live",
}


@dataclass
class EnrichedBlog:
    record: BlogRecord
    body: str = ""


@dataclass
class LinkVerdict:
    blog_name: str
    blog_link: str
    blog_date: str
    link_type: str
    link_label: str
    link_url: str
    score: float
    relevant: bool
    reason: str


def _tokenize(text: str) -> Set[str]:
    words = re.findall(r"[a-z0-9]+", (text or "").lower())
    return {w for w in words if len(w) > 2 and w not in STOP_WORDS}


def _slug_tokens(url: str) -> Set[str]:
    path = urlparse(url).path
    slug = path.rstrip("/").split("/")[-1]
    slug = re.sub(r"[%₹]", " ", slug)
    return _tokenize(slug.replace("-", " "))


def _normalize_label(label: str) -> str:
    clean = re.sub(r"[\[\]]", "", label or "").strip().lower()
    clean = re.sub(r"\s+", " ", clean)
    return clean


def _extract_body(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    el = soup.select_one(".entry-content, .post_content")
    return _clean_text(el.get_text()) if el else ""


def _extract_page_title(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    meta = soup.find("meta", property="og:title")
    if meta and meta.get("content"):
        return _clean_text(meta["content"])
    h1 = soup.find("h1", class_="entry-title") or soup.find("h1")
    if h1:
        return _clean_text(h1.get_text())
    return ""


def _blog_keywords(title: str, body: str, blog_url: str) -> Set[str]:
    kw = _tokenize(title)
    kw |= _tokenize(body[:2500])
    kw |= _slug_tokens(blog_url)
    return kw - DOMAIN_STOP


def _link_keywords(label: str, url: str, dest_title: str = "") -> Set[str]:
    clean_label = _normalize_label(label)
    if clean_label in GENERIC_LABELS or not clean_label:
        kw = _slug_tokens(url) | _tokenize(dest_title)
    else:
        kw = _tokenize(label) | _slug_tokens(url) | _tokenize(dest_title)
    return kw - DOMAIN_STOP


def score_relevance(
    blog_keywords: Set[str],
    link_keywords: Set[str],
) -> Tuple[float, str]:
    if not blog_keywords:
        return 0.0, "Could not extract blog topic keywords"
    if not link_keywords:
        return 0.0, "Could not extract link topic keywords"

    overlap = blog_keywords & link_keywords
    if overlap:
        score = len(overlap) / min(len(blog_keywords), max(len(link_keywords), 1), 10)
        score = min(score, 1.0)
        shared = ", ".join(sorted(overlap)[:6])
        return score, f"Shared topic words: {shared}"

    return 0.0, "No topical overlap between blog and link"


def scrape_enriched(session, listing: BlogListing) -> Optional[EnrichedBlog]:
    html = fetch(session, listing.url)
    if not html:
        return None
    rec = parse_blog_post(listing.url, html)
    if not rec.name and listing.title:
        rec.name = listing.title
    if not rec.date and listing.date:
        rec.date = listing.date
    return EnrichedBlog(record=rec, body=_extract_body(html))


def _dedupe_links(links: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
    seen: Set[str] = set()
    out: List[Tuple[str, str]] = []
    for label, url in links:
        key = url.rstrip("/").lower()
        if key in seen:
            continue
        seen.add(key)
        out.append((label, url))
    return out


def analyze_blog(
    blog: EnrichedBlog,
    threshold: float,
    title_cache: Dict[str, str],
    session,
    fetch_titles: bool,
) -> List[LinkVerdict]:
    rec = blog.record
    blog_kw = _blog_keywords(rec.name, blog.body, rec.link)
    verdicts: List[LinkVerdict] = []

    for link_type, links in (
        ("Internal", _dedupe_links(rec.internal_links)),
        ("External", _dedupe_links(rec.external_links)),
    ):
        for label, url in links:
            dest_title = ""
            if fetch_titles and _normalize_label(label) in GENERIC_LABELS:
                cache_key = url.rstrip("/").lower()
                if cache_key in title_cache:
                    dest_title = title_cache[cache_key]
                elif link_type == "Internal" and "asbl.in/blog/" in url:
                    html = fetch(session, url)
                    if html:
                        dest_title = _extract_page_title(html)
                        title_cache[cache_key] = dest_title

            link_kw = _link_keywords(label, url, dest_title)
            score, reason = score_relevance(blog_kw, link_kw)
            relevant = score >= threshold

            if not relevant and dest_title:
                reason = f"{reason}. Linked page: {dest_title[:80]}"

            verdicts.append(
                LinkVerdict(
                    blog_name=rec.name,
                    blog_link=rec.link,
                    blog_date=rec.date,
                    link_type=link_type,
                    link_label=label or url,
                    link_url=url,
                    score=round(score, 2),
                    relevant=relevant,
                    reason=reason,
                )
            )

    return verdicts


def write_report(
    all_verdicts: List[LinkVerdict],
    path: str,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    irrelevant = [v for v in all_verdicts if not v.relevant]

    blog_stats: Dict[str, dict] = {}
    for v in all_verdicts:
        key = v.blog_link
        if key not in blog_stats:
            blog_stats[key] = {
                "Blog Name": v.blog_name,
                "Blog Link": v.blog_link,
                "Blog Date": v.blog_date,
                "Total Internal Links": 0,
                "Total External Links": 0,
                "Irrelevant Internal": 0,
                "Irrelevant External": 0,
            }
        stat = blog_stats[key]
        if v.link_type == "Internal":
            stat["Total Internal Links"] += 1
            if not v.relevant:
                stat["Irrelevant Internal"] += 1
        else:
            stat["Total External Links"] += 1
            if not v.relevant:
                stat["Irrelevant External"] += 1

    summary_rows = []
    for stat in blog_stats.values():
        total_irr = stat["Irrelevant Internal"] + stat["Irrelevant External"]
        stat["Total Irrelevant Links"] = total_irr
        stat["Has Irrelevant Links"] = "Yes" if total_irr else "No"
        summary_rows.append(stat)

    summary_rows.sort(
        key=lambda r: (
            0 if r["Has Irrelevant Links"] == "Yes" else 1,
            -(r["Total Irrelevant Links"]),
            r["Blog Name"],
        )
    )

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C00000")
    ok_header_fill = PatternFill("solid", fgColor="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")
    irr_fill = PatternFill("solid", fgColor="FFC7CE")

    def style_header(ws, headers, fill=header_fill):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def set_link(cell, url: str, display: str):
        cell.value = display
        cell.hyperlink = url
        cell.font = link_font
        cell.alignment = wrap

    # Sheet 1: Irrelevant Links (primary deliverable)
    ws_irr = wb.active
    ws_irr.title = "Irrelevant Links"
    irr_headers = [
        "Blog Name",
        "Blog Link",
        "Blog Date",
        "Link Type",
        "Link Text",
        "Link URL",
        "Relevance Score",
        "Reason",
    ]
    style_header(ws_irr, irr_headers)
    for v in irrelevant:
        row = ws_irr.max_row + 1
        ws_irr.cell(row=row, column=1, value=v.blog_name).alignment = wrap
        set_link(ws_irr.cell(row=row, column=2), v.blog_link, v.blog_link)
        ws_irr.cell(row=row, column=3, value=v.blog_date).alignment = wrap
        ws_irr.cell(row=row, column=4, value=v.link_type).alignment = wrap
        ws_irr.cell(row=row, column=5, value=v.link_label).alignment = wrap
        set_link(ws_irr.cell(row=row, column=6), v.link_url, v.link_url)
        ws_irr.cell(row=row, column=7, value=v.score).alignment = wrap
        ws_irr.cell(row=row, column=8, value=v.reason).alignment = wrap
        for col in range(1, 9):
            ws_irr.cell(row=row, column=col).fill = irr_fill

    irr_widths = [45, 55, 14, 12, 40, 55, 16, 60]
    for i, w in enumerate(irr_widths, start=1):
        ws_irr.column_dimensions[get_column_letter(i)].width = w
    ws_irr.freeze_panes = "A2"
    if ws_irr.max_row > 1:
        ws_irr.auto_filter.ref = ws_irr.dimensions

    # Sheet 2: Blog Summary
    ws_sum = wb.create_sheet("Blog Summary")
    sum_headers = [
        "Blog Name",
        "Blog Link",
        "Blog Date",
        "Total Internal Links",
        "Total External Links",
        "Irrelevant Internal",
        "Irrelevant External",
        "Total Irrelevant Links",
        "Has Irrelevant Links",
    ]
    style_header(ws_sum, sum_headers, ok_header_fill)
    for stat in summary_rows:
        row = ws_sum.max_row + 1
        ws_sum.cell(row=row, column=1, value=stat["Blog Name"]).alignment = wrap
        set_link(ws_sum.cell(row=row, column=2), stat["Blog Link"], stat["Blog Link"])
        ws_sum.cell(row=row, column=3, value=stat["Blog Date"]).alignment = wrap
        for col, key in enumerate(
            [
                "Total Internal Links",
                "Total External Links",
                "Irrelevant Internal",
                "Irrelevant External",
                "Total Irrelevant Links",
                "Has Irrelevant Links",
            ],
            start=4,
        ):
            ws_sum.cell(row=row, column=col, value=stat[key]).alignment = wrap
            if stat["Has Irrelevant Links"] == "Yes" and col >= 6:
                ws_sum.cell(row=row, column=col).fill = irr_fill

    sum_widths = [45, 55, 14, 18, 18, 18, 18, 20, 18]
    for i, w in enumerate(sum_widths, start=1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = ws_sum.dimensions

    # Sheet 3: All Links (full audit)
    ws_all = wb.create_sheet("All Links")
    all_headers = irr_headers + ["Verdict"]
    style_header(ws_all, all_headers, ok_header_fill)
    for v in all_verdicts:
        row = ws_all.max_row + 1
        ws_all.cell(row=row, column=1, value=v.blog_name).alignment = wrap
        set_link(ws_all.cell(row=row, column=2), v.blog_link, v.blog_link)
        ws_all.cell(row=row, column=3, value=v.blog_date).alignment = wrap
        ws_all.cell(row=row, column=4, value=v.link_type).alignment = wrap
        ws_all.cell(row=row, column=5, value=v.link_label).alignment = wrap
        set_link(ws_all.cell(row=row, column=6), v.link_url, v.link_url)
        ws_all.cell(row=row, column=7, value=v.score).alignment = wrap
        ws_all.cell(row=row, column=8, value=v.reason).alignment = wrap
        verdict = "Relevant" if v.relevant else "Irrelevant"
        cell = ws_all.cell(row=row, column=9, value=verdict)
        cell.alignment = wrap
        if not v.relevant:
            for col in range(1, 10):
                ws_all.cell(row=row, column=col).fill = irr_fill

    all_widths = irr_widths + [12]
    for i, w in enumerate(all_widths, start=1):
        ws_all.column_dimensions[get_column_letter(i)].width = w
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = ws_all.dimensions

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Find internal/external blog links that are not relevant to the article.",
    )
    parser.add_argument("--start", type=int, default=1, help="First listing page. Default: 1")
    parser.add_argument("--end", type=int, default=43, help="Last listing page. Default: 43")
    parser.add_argument("--workers", type=int, default=8, help="Parallel workers. Default: 8")
    parser.add_argument("--out", default="irrelevant_links.xlsx", help="Output Excel path.")
    parser.add_argument("--limit", type=int, default=0, help="Cap posts (debugging).")
    parser.add_argument(
        "--threshold",
        type=float,
        default=0.15,
        help="Relevance score cutoff (0-1). Below = irrelevant. Default: 0.15",
    )
    parser.add_argument(
        "--fetch-titles",
        action="store_true",
        help="Fetch linked page titles for generic [Source] anchors (slower, more accurate).",
    )
    args = parser.parse_args()

    session = make_session()
    listings = collect_listings(session, args.start, args.end)
    if args.limit:
        listings = listings[: args.limit]

    print(f"Analyzing {len(listings)} blog posts...", file=sys.stderr)

    enriched: List[EnrichedBlog] = []
    failed: List[str] = []

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {
            pool.submit(scrape_enriched, make_session(), entry): entry
            for entry in listings
        }
        for fut in tqdm(as_completed(futures), total=len(futures), desc="Scraping posts"):
            entry = futures[fut]
            try:
                item = fut.result()
            except Exception as exc:  # noqa: BLE001
                print(f"  ! Error: {entry.url}: {exc}", file=sys.stderr)
                failed.append(entry.url)
                continue
            if item is None:
                failed.append(entry.url)
                continue
            enriched.append(item)

    order = {entry.url: i for i, entry in enumerate(listings)}
    enriched.sort(key=lambda b: order.get(b.record.link, 1_000_000))

    title_cache: Dict[str, str] = {}
    all_verdicts: List[LinkVerdict] = []

    for blog in tqdm(enriched, desc="Scoring links"):
        verdicts = analyze_blog(
            blog,
            threshold=args.threshold,
            title_cache=title_cache,
            session=session,
            fetch_titles=args.fetch_titles,
        )
        all_verdicts.extend(verdicts)

    write_report(all_verdicts, args.out)

    irrelevant_count = sum(1 for v in all_verdicts if not v.relevant)
    blogs_with_issues = len(
        {v.blog_link for v in all_verdicts if not v.relevant}
    )
    blogs_with_links = len({v.blog_link for v in all_verdicts})

    print(
        f"\nSaved report -> {args.out}\n"
        f"  Posts scraped:           {len(enriched)} ({len(failed)} failed)\n"
        f"  Posts with styled links: {blogs_with_links}\n"
        f"  Total links scored:      {len(all_verdicts)}\n"
        f"  Irrelevant links:        {irrelevant_count}\n"
        f"  Blogs with issues:       {blogs_with_issues}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
